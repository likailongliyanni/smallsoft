"""Create a clean product-catalog workbook from heterogeneous local sheets."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT_HEADERS = ["商品名称", "品牌", "型号", "编码", "条码", "电商链接", "销售价格", "代发价格"]
SHEET_BRANDS = {
    "SKG": "SKG", "有色": "有色/yoose", "徕芬": "徕芬", "左点": "Zdeer/左点",
    "兰士顿": "兰士顿", "小胖爪": "小胖爪", "cleer": "Cleer/可丽尔",
    "索哈曼": "索哈曼SOHOMAN", "艾贝丽": "Abereve/艾贝丽", "西屋": "西屋",
}
FIELD_ALIASES = {
    "name": ["商品名称", "产品全称", "产品名称", "名称", "sku名称", "商品型号"],
    "brand": ["品牌"],
    "model": ["产品型号", "型号/规格", "型号", "商品型号", "规格"],
    "code": ["产品编码", "商品编码", "物料", "货号", "sku编码"],
    "barcode": ["商品条码", "产品条码", "产品条形码", "商品69码", "69码", "颜色+条码", "69"],
    "link": ["电商链接", "京东/天猫旗舰店链接", "天猫/京东旗舰店链接", "京东旗舰链接",
             "官旗店链接", "京东/天猫链接", "京东链接", "链接", "京东"],
    "sale": ["销售价格", "标准零售价", "官网标价（人民币）", "京东价", "日常价",
             "网价或市场价", "京东价格", "旗舰店价格", "零售价"],
    "dropship": ["代发价格", "一件代发价（含13%税含普通快递运费）",
                 "客户一件代发价大单项目报价单独沟通", "一件代发（含税运）", "一件代发", "代发价", "代发"],
}


def normalize(value: Any) -> str:
    return re.sub(r"[\s\n\r（）()【】\[\]_*·:：、，,/]+", "", str(value or "")).lower()


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("_x000B_", " ")).strip()


def identifier(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return cell_text(value)


def price_value(value: Any) -> Any:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    raw = cell_text(value)
    cleaned = re.sub(r"[￥¥$,，]", "", raw)
    if re.fullmatch(r"-?\d+(?:\.\d+)?", cleaned):
        return float(cleaned) if "." in cleaned else int(cleaned)
    return raw


def header_map(headers: list[Any]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(headers):
        key = normalize(value)
        if key and key not in result:
            result[key] = index
    return result


def get_value(row: list[Any], mapping: dict[str, int], aliases: list[str]) -> Any:
    for alias in aliases:
        index = mapping.get(normalize(alias))
        if index is not None and index < len(row) and cell_text(row[index]):
            return row[index]
    return ""


def detect_header_row(rows: list[list[Any]]) -> int:
    known = {normalize(alias) for aliases in FIELD_ALIASES.values() for alias in aliases}
    known.update({normalize("产品参数"), normalize("物料")})
    best = (0, 0)
    for index, row in enumerate(rows[:20]):
        score = sum(normalize(value) in known for value in row if cell_text(value))
        if score > best[1]:
            best = (index, score)
    return best[0]


def model_from_parameters(value: Any) -> str:
    match = re.search(r"型号\s*[:：/]\s*([A-Za-z0-9][A-Za-z0-9._-]{1,30})", cell_text(value), re.I)
    return match.group(1) if match else ""


def models_from_name(value: Any) -> str:
    matches = re.findall(r"\b(?:W[A-Z0-9]+(?:-[A-Z0-9]+)+|[A-Z]{1,5}\d[A-Z0-9-]*)\b", cell_text(value), re.I)
    unique = []
    for item in matches:
        cleaned = item.rstrip("-_")
        if cleaned and cleaned.lower() not in {value.lower() for value in unique}:
            unique.append(cleaned)
    return " / ".join(unique)


def clean_link(value: Any) -> str:
    result = cell_text(value)
    if result.startswith("ttps://"):
        result = "h" + result
    return result


def _sheet_rows(sheet) -> list[list[Any]]:
    max_column = min(max(1, int(sheet.max_column or 1)), 100)
    max_row = min(max(1, int(sheet.max_row or 1)), 50000)
    return [list(row) for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column, values_only=True)]


def extract_product_rows(source_path: Path) -> tuple[list[list[Any]], dict[str, int]]:
    workbook = load_workbook(source_path, read_only=True, data_only=False, keep_links=False)
    try:
        raw_sheets = [(sheet.title, _sheet_rows(sheet)) for sheet in workbook.worksheets]
    finally:
        workbook.close()

    skg_codes: dict[str, str] = {}
    for sheet_name, rows in raw_sheets:
        if sheet_name != "Sheet1" or not rows:
            continue
        header_index = detect_header_row(rows)
        mapping = header_map(rows[header_index])
        for row in rows[header_index + 1:]:
            barcode = identifier(get_value(row, mapping, ["69码"]))
            code = identifier(get_value(row, mapping, ["物料", "货号"]))
            if barcode and code:
                skg_codes[barcode] = code

    extracted: list[list[Any]] = []
    for sheet_name, rows in raw_sheets:
        if sheet_name == "Sheet1" or len(rows) < 2:
            continue
        header_index = detect_header_row(rows)
        mapping = header_map(rows[header_index])
        for row in rows[header_index + 1:]:
            name = get_value(row, mapping, FIELD_ALIASES["name"])
            brand = get_value(row, mapping, FIELD_ALIASES["brand"]) or SHEET_BRANDS.get(sheet_name, sheet_name)
            model = get_value(row, mapping, FIELD_ALIASES["model"])
            code = get_value(row, mapping, FIELD_ALIASES["code"])
            barcode = get_value(row, mapping, FIELD_ALIASES["barcode"])
            link = get_value(row, mapping, FIELD_ALIASES["link"])
            sale = get_value(row, mapping, FIELD_ALIASES["sale"])
            dropship = get_value(row, mapping, FIELD_ALIASES["dropship"])

            if sheet_name == "有色":
                name = get_value(row, mapping, ["sku名称", "产品名称"])
                model = get_value(row, mapping, ["产品名称"])
            elif sheet_name == "左点":
                name = get_value(row, mapping, ["商品型号"])
                model = model_from_parameters(get_value(row, mapping, ["产品参数"])) or cell_text(name)
            elif sheet_name == "cleer":
                name = get_value(row, mapping, ["商品型号"])
                model = name
            elif sheet_name == "西屋":
                model = models_from_name(name)

            clean_name = cell_text(name)
            if not clean_name or clean_name in {"合计", "总计", "备注"}:
                continue
            clean_barcode = identifier(barcode)
            if sheet_name == "SKG" and not cell_text(code):
                code = skg_codes.get(clean_barcode, "")
            extracted.append([
                clean_name, cell_text(brand), cell_text(model), identifier(code), clean_barcode,
                clean_link(link), price_value(sale), price_value(dropship),
            ])

    unique: list[list[Any]] = []
    seen: set[str] = set()
    for row in extracted:
        key = f"barcode:{row[4]}" if row[4] else f"code:{row[3]}" if row[3] else f"fallback:{row[1]}|{row[2]}|{row[0]}"
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    missing = {header: sum(row[index] in {"", None} for row in unique) for index, header in enumerate(OUTPUT_HEADERS)}
    return unique, missing


def create_product_catalog(source_path: Path, output_path: Path) -> dict[str, Any]:
    rows, missing = extract_product_rows(source_path)
    if not rows:
        raise RuntimeError("没有识别到可整理的商品明细，请确认各工作表包含商品表头")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "商品汇总"
    sheet.append(OUTPUT_HEADERS)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = f"A1:H{len(rows) + 1}"

    header_fill = PatternFill("solid", fgColor="167A50")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30
    widths = {"A": 42, "B": 18, "C": 22, "D": 18, "E": 24, "F": 48, "G": 14, "H": 14}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row_index in range(2, len(rows) + 2):
        sheet.row_dimensions[row_index].height = 22
        sheet.cell(row_index, 1).alignment = Alignment(vertical="center", wrap_text=True)
        for column_index in (4, 5):
            sheet.cell(row_index, column_index).number_format = "@"
        for column_index in (7, 8):
            sheet.cell(row_index, column_index).number_format = '¥#,##0.00'
        link_cell = sheet.cell(row_index, 6)
        if isinstance(link_cell.value, str) and link_cell.value.startswith(("http://", "https://")):
            link_cell.hyperlink = link_cell.value.split()[0]
            link_cell.font = Font(name="Microsoft YaHei", size=10, color="1D4ED8", underline="single")
    for row in sheet.iter_rows(min_row=2, max_row=len(rows) + 1, min_col=1, max_col=8):
        for cell in row:
            if cell.column != 6:
                cell.font = Font(name="Microsoft YaHei", size=10, color="24332B")
            if cell.alignment == Alignment():
                cell.alignment = Alignment(vertical="center")

    table = Table(displayName="ProductCatalog", ref=f"A1:H{len(rows) + 1}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False,
                                         showRowStripes=True, showColumnStripes=False)
    sheet.add_table(table)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {"rows": len(rows), "missing": missing, "file_path": str(output_path), "file_name": output_path.name}
