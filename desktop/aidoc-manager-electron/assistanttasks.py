"""Local, read-only spreadsheet task inbox for the AI document secretary."""

from __future__ import annotations

import csv
import hashlib
import json
import shutil
import sqlite3
import threading
import uuid
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import sheetorganizer


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".csv", ".tsv"}
_RUNNING: set[str] = set()
_RUNNING_LOCK = threading.Lock()


def now_text() -> str:
    return datetime.now().isoformat(timespec="seconds")


def ensure_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS assistant_tasks (
          id TEXT PRIMARY KEY,
          title TEXT NOT NULL,
          instruction TEXT NOT NULL DEFAULT '',
          status TEXT NOT NULL DEFAULT 'received',
          summary TEXT NOT NULL DEFAULT '',
          analysis_json TEXT NOT NULL DEFAULT '{}',
          error TEXT NOT NULL DEFAULT '',
          conversation_id TEXT NOT NULL DEFAULT '',
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL,
          completed_at TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS assistant_task_files (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          task_id TEXT NOT NULL,
          role TEXT NOT NULL DEFAULT 'original',
          original_name TEXT NOT NULL,
          original_path TEXT NOT NULL DEFAULT '',
          stored_path TEXT NOT NULL,
          sha256 TEXT NOT NULL,
          extension TEXT NOT NULL,
          size_bytes INTEGER NOT NULL DEFAULT 0,
          created_at TEXT NOT NULL,
          FOREIGN KEY(task_id) REFERENCES assistant_tasks(id) ON DELETE CASCADE
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_assistant_tasks_status ON assistant_tasks(status, updated_at)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_assistant_task_files_task ON assistant_task_files(task_id, id)")
    conn.commit()


def _connect(db_file: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(str(db_file), timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout = 30000")
    ensure_tables(conn)
    return conn


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _safe_title(instruction: str, names: list[str]) -> str:
    first_line = instruction.strip().splitlines()[0][:60] if instruction.strip() else ""
    if first_line:
        return first_line
    if len(names) == 1:
        return f"梳理 {names[0]}"
    return f"梳理 {len(names)} 份表格"


def create_task(
    db_file: Path,
    library_root: Path,
    paths: list[str],
    instruction: str,
    conversation_id: str = "",
) -> dict[str, Any]:
    sources: list[Path] = []
    for value in paths:
        path = Path(str(value or "")).expanduser().resolve()
        if not path.is_file():
            raise RuntimeError(f"找不到表格文件：{path.name or value}")
        if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            raise RuntimeError(f"暂不支持 {path.suffix or '未知格式'}，请使用 xlsx、xlsm、csv 或 tsv")
        sources.append(path)
    if not sources:
        raise RuntimeError("请先选择要托管的表格")

    task_id = datetime.now().strftime("%Y%m%d%H%M%S") + "-" + uuid.uuid4().hex[:8]
    stamp = now_text()
    assistant_root = library_root / ".assistant"
    objects_dir = assistant_root / "objects"
    task_dir = assistant_root / "tasks" / task_id
    task_dir.mkdir(parents=True, exist_ok=True)
    objects_dir.mkdir(parents=True, exist_ok=True)

    file_rows: list[dict[str, Any]] = []
    for source in sources:
        sha = _sha256(source)
        extension = source.suffix.lower()
        object_dir = objects_dir / sha[:2]
        object_dir.mkdir(parents=True, exist_ok=True)
        stored = object_dir / f"{sha}{extension}"
        if not stored.exists():
            shutil.copy2(source, stored)
        file_rows.append({
            "original_name": source.name,
            "original_path": str(source),
            "stored_path": str(stored),
            "sha256": sha,
            "extension": extension,
            "size_bytes": source.stat().st_size,
        })

    conn = _connect(db_file)
    try:
        conn.execute(
            """INSERT INTO assistant_tasks
               (id, title, instruction, status, conversation_id, created_at, updated_at)
               VALUES (?, ?, ?, 'received', ?, ?, ?)""",
            (task_id, _safe_title(instruction, [x.name for x in sources]), instruction.strip(), conversation_id, stamp, stamp),
        )
        conn.executemany(
            """INSERT INTO assistant_task_files
               (task_id, role, original_name, original_path, stored_path, sha256, extension, size_bytes, created_at)
               VALUES (?, 'original', ?, ?, ?, ?, ?, ?, ?)""",
            [
                (task_id, row["original_name"], row["original_path"], row["stored_path"], row["sha256"],
                 row["extension"], row["size_bytes"], stamp)
                for row in file_rows
            ],
        )
        conn.commit()
        task = get_task(conn, task_id)
    finally:
        conn.close()
    schedule_analysis(db_file, task_id)
    return task


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value).strip()


def _profile_rows(name: str, rows: list[list[Any]], formulas: int = 0, errors: int = 0) -> dict[str, Any]:
    nonempty = [row for row in rows if any(_cell_text(value) for value in row)]
    width = max((len(row) for row in nonempty), default=0)
    header_index = 0
    for index, row in enumerate(rows[:20]):
        values = [_cell_text(value) for value in row]
        if sum(bool(value) for value in values) >= 2:
            header_index = index
            break
    headers = [_cell_text(value) or f"第{index + 1}列" for index, value in enumerate(rows[header_index] if rows else [])]
    headers = headers[:30]
    samples = []
    for row in rows[header_index + 1:]:
        values = [_cell_text(value) for value in row[:30]]
        if any(values):
            samples.append(values)
        if len(samples) >= 5:
            break
    normalized = [tuple(_cell_text(value) for value in row) for row in nonempty[header_index + 1:50001]]
    duplicates = sum(count - 1 for count in Counter(normalized).values() if count > 1)
    blank_rows = max(0, len(rows) - len(nonempty))
    return {
        "sheet": name,
        "rows": len(rows),
        "nonempty_rows": len(nonempty),
        "columns": width,
        "header_row": header_index + 1 if rows else 0,
        "headers": headers,
        "sample_rows": samples,
        "blank_rows": blank_rows,
        "duplicate_rows": duplicates,
        "formula_cells": formulas,
        "error_cells": errors,
    }


def _profile_excel(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    sheets: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            rows: list[list[Any]] = []
            formulas = 0
            errors = 0
            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                values = list(row)
                formulas += sum(isinstance(value, str) and value.startswith("=") for value in values)
                errors += sum(isinstance(value, str) and value.startswith("#") for value in values)
                rows.append(values)
                if row_index >= 49999:
                    break
            item = _profile_rows(sheet.title, rows, formulas, errors)
            item["truncated"] = sheet.max_row > 50000
            sheets.append(item)
    finally:
        workbook.close()
    return sheets


def _read_delimited(path: Path, delimiter: str) -> list[list[str]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return [row for index, row in enumerate(csv.reader(handle, delimiter=delimiter)) if index < 50000]
        except UnicodeDecodeError as exc:
            last_error = exc
    raise RuntimeError(f"无法识别文本编码：{last_error}")


def analyze_file(path: Path, extension: str) -> list[dict[str, Any]]:
    if extension in {".xlsx", ".xlsm"}:
        return _profile_excel(path)
    delimiter = "\t" if extension == ".tsv" else ","
    return [_profile_rows(path.stem, _read_delimited(path, delimiter))]


def _build_summary(files: list[dict[str, Any]]) -> str:
    sheet_count = sum(len(item["sheets"]) for item in files)
    rows = sum(sheet["nonempty_rows"] for item in files for sheet in item["sheets"])
    duplicates = sum(sheet["duplicate_rows"] for item in files for sheet in item["sheets"])
    blanks = sum(sheet["blank_rows"] for item in files for sheet in item["sheets"])
    errors = sum(sheet["error_cells"] for item in files for sheet in item["sheets"])
    lines = [f"已完成只读梳理：{len(files)} 个文件，{sheet_count} 个工作表，约 {rows} 行有效数据。"]
    details = []
    if duplicates:
        details.append(f"发现 {duplicates} 行重复数据")
    if blanks:
        details.append(f"发现 {blanks} 行空行")
    if errors:
        details.append(f"发现 {errors} 个疑似公式错误单元格")
    lines.append("；".join(details) + "。" if details else "未发现明显的重复行、空行或公式错误。")
    for item in files[:5]:
        sheet_text = "、".join(
            f"{sheet['sheet']}（{sheet['nonempty_rows']}行×{sheet['columns']}列）" for sheet in item["sheets"][:8]
        )
        lines.append(f"• {item['name']}：{sheet_text}")
    lines.append("原表未修改。需要时可以继续让我核对字段、找异常或整理成正式版本。")
    return "\n".join(lines)


def _analyze_task(db_file: Path, task_id: str) -> None:
    conn = _connect(db_file)
    try:
        conn.execute("UPDATE assistant_tasks SET status = 'processing', error = '', updated_at = ? WHERE id = ?", (now_text(), task_id))
        conn.commit()
        rows = conn.execute(
            "SELECT * FROM assistant_task_files WHERE task_id = ? AND role = 'original' ORDER BY id", (task_id,)
        ).fetchall()
    finally:
        conn.close()

    try:
        files = []
        for row in rows:
            path = Path(str(row["stored_path"]))
            files.append({"name": str(row["original_name"]), "sheets": analyze_file(path, str(row["extension"]))})
        summary = _build_summary(files)
        if rows:
            assistant_root = Path(str(rows[0]["stored_path"])).parents[2]
            result_dir = assistant_root / "tasks" / task_id
            result_dir.mkdir(parents=True, exist_ok=True)
            (result_dir / "梳理结果.txt").write_text(summary, encoding="utf-8")
        conn = _connect(db_file)
        try:
            stamp = now_text()
            conn.execute(
                """UPDATE assistant_tasks SET status = 'ready', summary = ?, analysis_json = ?,
                   error = '', updated_at = ?, completed_at = ? WHERE id = ?""",
                (summary, json.dumps({"files": files}, ensure_ascii=False), stamp, stamp, task_id),
            )
            conn.commit()
        finally:
            conn.close()
    except Exception as exc:
        conn = _connect(db_file)
        try:
            conn.execute(
                "UPDATE assistant_tasks SET status = 'error', error = ?, updated_at = ? WHERE id = ?",
                (str(exc)[:1000], now_text(), task_id),
            )
            conn.commit()
        finally:
            conn.close()
    finally:
        with _RUNNING_LOCK:
            _RUNNING.discard(task_id)


def schedule_analysis(db_file: Path, task_id: str) -> None:
    with _RUNNING_LOCK:
        if task_id in _RUNNING:
            return
        _RUNNING.add(task_id)
    threading.Thread(target=_analyze_task, args=(db_file, task_id), name=f"sheet-task-{task_id}", daemon=True).start()


def resume_pending(db_file: Path) -> None:
    conn = _connect(db_file)
    try:
        ids = [str(row[0]) for row in conn.execute(
            "SELECT id FROM assistant_tasks WHERE status IN ('received', 'processing') ORDER BY created_at LIMIT 20"
        ).fetchall()]
    finally:
        conn.close()
    for task_id in ids:
        schedule_analysis(db_file, task_id)


def _task_payload(conn: sqlite3.Connection, row: sqlite3.Row, include_analysis: bool = False) -> dict[str, Any]:
    files = [dict(item) for item in conn.execute(
        "SELECT id, role, original_name, original_path, stored_path, extension, size_bytes FROM assistant_task_files WHERE task_id = ? ORDER BY id",
        (row["id"],),
    ).fetchall()]
    payload = {
        "id": str(row["id"]),
        "title": str(row["title"]),
        "instruction": str(row["instruction"]),
        "status": str(row["status"]),
        "summary": str(row["summary"]),
        "error": str(row["error"]),
        "conversation_id": str(row["conversation_id"]),
        "created_at": str(row["created_at"]),
        "updated_at": str(row["updated_at"]),
        "completed_at": str(row["completed_at"] or ""),
        "files": files,
        "task_dir": str(Path(files[0]["stored_path"]).parents[2] / "tasks" / str(row["id"])) if files else "",
    }
    if include_analysis:
        try:
            payload["analysis"] = json.loads(str(row["analysis_json"] or "{}"))
        except json.JSONDecodeError:
            payload["analysis"] = {}
    return payload


def get_task(conn: sqlite3.Connection, task_id: str) -> dict[str, Any]:
    row = conn.execute("SELECT * FROM assistant_tasks WHERE id = ?", (task_id,)).fetchone()
    if row is None:
        raise RuntimeError("托管事项不存在")
    return _task_payload(conn, row, include_analysis=True)


def list_tasks(conn: sqlite3.Connection, limit: int = 100) -> list[dict[str, Any]]:
    rows = conn.execute(
        "SELECT * FROM assistant_tasks ORDER BY created_at DESC LIMIT ?", (max(1, min(limit, 300)),)
    ).fetchall()
    return [_task_payload(conn, row) for row in rows]


def recent_context(conn: sqlite3.Connection, limit: int = 12) -> list[dict[str, Any]]:
    rows = conn.execute(
        "SELECT * FROM assistant_tasks ORDER BY updated_at DESC LIMIT ?", (max(1, min(limit, 20)),)
    ).fetchall()
    result = []
    for row in rows:
        item = _task_payload(conn, row)
        result.append({
            "id": item["id"],
            "title": item["title"],
            "instruction": item["instruction"][:500],
            "status": item["status"],
            "summary": item["summary"][:3000],
            "files": [file["original_name"] for file in item["files"]],
            "updated_at": item["updated_at"],
        })
    return result


def archive_task(conn: sqlite3.Connection, task_id: str) -> dict[str, Any]:
    conn.execute("UPDATE assistant_tasks SET status = 'archived', updated_at = ? WHERE id = ?", (now_text(), task_id))
    conn.commit()
    return get_task(conn, task_id)


def organize_product_catalog(db_file: Path, task_id: str) -> dict[str, Any]:
    """Generate a new, clean product workbook; never modify the managed original."""
    conn = _connect(db_file)
    try:
        task = get_task(conn, task_id)
        source_file = next(
            (item for item in task["files"] if item["role"] == "original" and item["extension"] in {".xlsx", ".xlsm"}),
            None,
        )
        if source_file is None:
            raise RuntimeError("当前托管事项没有可整理的 Excel 文件")
        conn.execute("UPDATE assistant_tasks SET status = 'processing', error = '', updated_at = ? WHERE id = ?", (now_text(), task_id))
        conn.commit()
    finally:
        conn.close()

    source_path = Path(str(source_file["stored_path"]))
    assistant_root = source_path.parents[2]
    output_dir = assistant_root / "tasks" / task_id / "outputs"
    safe_stem = re.sub(r"[\\/:*?\"<>|]+", "_", Path(str(source_file["original_name"])).stem).strip() or "托管表格"
    output_path = output_dir / f"{safe_stem}_商品字段整理版.xlsx"
    try:
        result = sheetorganizer.create_product_catalog(source_path, output_path)
        result_sha = _sha256(output_path)
        summary = (
            f"已按要求生成整理版 Excel：从多个工作表提取商品名称、品牌、型号、编码、条码、"
            f"电商链接、销售价格、代发价格，共 {result['rows']} 条商品记录。原表未修改。"
        )
        conn = _connect(db_file)
        try:
            row = conn.execute("SELECT analysis_json FROM assistant_tasks WHERE id = ?", (task_id,)).fetchone()
            try:
                analysis = json.loads(str(row["analysis_json"] or "{}")) if row else {}
            except json.JSONDecodeError:
                analysis = {}
            analysis["organization"] = {"type": "product_catalog", **result}
            conn.execute("DELETE FROM assistant_task_files WHERE task_id = ? AND role = 'output'", (task_id,))
            conn.execute(
                """INSERT INTO assistant_task_files
                   (task_id, role, original_name, original_path, stored_path, sha256, extension, size_bytes, created_at)
                   VALUES (?, 'output', ?, '', ?, ?, '.xlsx', ?, ?)""",
                (task_id, output_path.name, str(output_path), result_sha, output_path.stat().st_size, now_text()),
            )
            stamp = now_text()
            conn.execute(
                """UPDATE assistant_tasks SET status = 'complete', summary = ?, analysis_json = ?, error = '',
                   updated_at = ?, completed_at = ? WHERE id = ?""",
                (summary, json.dumps(analysis, ensure_ascii=False), stamp, stamp, task_id),
            )
            conn.commit()
            task = get_task(conn, task_id)
        finally:
            conn.close()
        return {**result, "dir": str(output_dir), "task": task}
    except Exception as exc:
        conn = _connect(db_file)
        try:
            conn.execute(
                "UPDATE assistant_tasks SET status = 'error', error = ?, updated_at = ? WHERE id = ?",
                (str(exc)[:1000], now_text(), task_id),
            )
            conn.commit()
        finally:
            conn.close()
        raise
