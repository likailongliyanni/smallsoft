import sqlite3
import tempfile
import time
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

import assistanttasks
import sheetorganizer


class AssistantTaskTests(unittest.TestCase):
    def test_spreadsheet_is_copied_and_analyzed_without_changing_source(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "客户对账表.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "六月对账"
            sheet.append(["客户", "金额", "备注"])
            sheet.append(["甲公司", 1200, "已核对"])
            sheet.append(["甲公司", 1200, "已核对"])
            sheet.append([None, None, None])
            workbook.save(source)
            original_hash = assistanttasks._sha256(source)

            db_file = root / "aidoc.db"
            task = assistanttasks.create_task(db_file, root, [str(source)], "客户六月对账表，先帮我梳理")
            self.assertEqual(task["status"], "received")
            self.assertNotEqual(task["files"][0]["stored_path"], str(source))

            deadline = time.time() + 5
            current = task
            while time.time() < deadline:
                conn = sqlite3.connect(db_file)
                conn.row_factory = sqlite3.Row
                try:
                    current = assistanttasks.get_task(conn, task["id"])
                finally:
                    conn.close()
                if current["status"] in {"ready", "error"}:
                    break
                time.sleep(0.05)

            self.assertEqual(current["status"], "ready", current.get("error"))
            self.assertIn("1 个工作表", current["summary"])
            self.assertIn("重复数据", current["summary"])
            self.assertEqual(assistanttasks._sha256(source), original_hash)
            self.assertEqual(current["analysis"]["files"][0]["sheets"][0]["headers"][:3], ["客户", "金额", "备注"])

    def test_rejects_non_spreadsheet_file(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "说明.txt"
            source.write_text("hello", encoding="utf-8")
            with self.assertRaisesRegex(RuntimeError, "暂不支持"):
                assistanttasks.create_task(root / "aidoc.db", root, [str(source)], "帮我看看")

    def test_product_catalog_organizer_merges_sheets_and_enriches_skg_code(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "货盘.xlsx"
            workbook = Workbook()
            skg = workbook.active
            skg.title = "SKG"
            skg.append(["产品名称", "型号", "69码", "标准零售价", "代发", "链接"])
            skg.append(["颈椎按摩仪", "K1", 6944527434198, 499, 170, "https://example.com/k1"])
            pet = workbook.create_sheet("小胖爪")
            pet.append(["产品名称", "产品编码", "规格", "商品条码", "官旗店链接", "日常价", "代发价"])
            pet.append(["猫粮", "CH001", "1.5kg", 6974917750287, "https://example.com/cat", 129, 70])
            tech = workbook.create_sheet("Sheet1")
            tech.append(["物料", "名称", "69码"])
            tech.append(["1105070007", "颈椎按摩仪", 6944527434198])
            workbook.save(source)

            output = root / "整理版.xlsx"
            result = sheetorganizer.create_product_catalog(source, output)
            self.assertEqual(2, result["rows"])
            organized = load_workbook(output, data_only=True)
            try:
                sheet = organized["商品汇总"]
                self.assertEqual(sheetorganizer.OUTPUT_HEADERS, [cell.value for cell in sheet[1]])
                self.assertEqual("1105070007", sheet["D2"].value)
                self.assertEqual("CH001", sheet["D3"].value)
                self.assertEqual("6974917750287", sheet["E3"].value)
            finally:
                organized.close()


if __name__ == "__main__":
    unittest.main()
