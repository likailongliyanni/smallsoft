"""Local backend for Haobanfa AI document manager.

The desktop app never uploads original files in this first architecture.
It scans local folders, counts billable pages, stores metadata in a local
SQLite library, and sends only device/quota requests to tools.haobanfa.
Future OCR/AI extraction commands should send text, not the source file.
"""

from __future__ import annotations

import base64
import csv
import hashlib
import hmac
import io
import json
import os
import queue
import re
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import threading
import time
import urllib.error
import urllib.request
import uuid
import zipfile
from concurrent.futures import CancelledError, Future, ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import docintel
import assistanttasks
import contractgen
import documentgen
import knowledge
import templatepool

APP_CODE = "AIDOC"
APP_NAME = "ai-doc"
APP_VERSION = "1.0.3"
FREE_POINTS = 30
DEFAULT_SERVER_URL = "https://tools.haobanfa.online"

SUPPORTED_EXTENSIONS = {
    ".pdf",
    ".jpg",
    ".jpeg",
    ".png",
    ".webp",
    ".bmp",
    ".tif",
    ".tiff",
    ".docx",
    ".xlsx",
    ".xlsm",
    ".txt",
    ".csv",
    ".md",
    ".markdown",
    ".json",
    ".xml",
    ".rtf",
    ".log",
}

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff"}
TEXT_EXTENSIONS = {".txt", ".csv", ".md", ".markdown", ".json", ".xml", ".rtf", ".log"}
OFFICE_EXTENSIONS = {".docx", ".xlsx", ".xlsm"}

APP_DIR = Path(os.environ.get("APPDATA") or Path.home() / "AppData" / "Roaming") / "HaobanfaAIDoc"
CONFIG_PATH = APP_DIR / "config.json"

STATE: dict[str, Any] = {"token": "", "unlocked": False}
CONFIG_LOCK = threading.Lock()
EMIT_LOCK = threading.Lock()
RECOGNIZE_CONTEXT = threading.local()
RECOGNITION_JOBS: dict[str, dict[str, Any]] = {}
RECOGNITION_JOBS_LOCK = threading.Lock()
KNOWLEDGE_QUEUE: queue.Queue[tuple[int, bool]] = queue.Queue()
KNOWLEDGE_QUEUE_LOCK = threading.Lock()
KNOWLEDGE_QUEUED: set[tuple[int, bool]] = set()
KNOWLEDGE_WORKER: threading.Thread | None = None


def emit(payload: dict[str, Any]) -> None:
    # PyInstaller 的无控制台进程在部分中文 Windows 上仍会把文本管道设为 GBK，
    # 即使父进程传了 PYTHONIOENCODING。直接写二进制 UTF-8，避免文件名、类型和
    # 服务端错误消息在 Electron 里变成乱码。
    raw = (json.dumps(payload, ensure_ascii=False) + "\n").encode("utf-8")
    # 批量识别时多个工作线程会同时推送进度，必须保证每条 JSON 完整写入。
    with EMIT_LOCK:
        binary = getattr(sys.stdout, "buffer", None)
        if binary is not None:
            binary.write(raw)
            binary.flush()
            return
        sys.stdout.write(raw.decode("utf-8"))
        sys.stdout.flush()


def load_config() -> dict[str, Any]:
    APP_DIR.mkdir(parents=True, exist_ok=True)
    if not CONFIG_PATH.exists():
        return {
            "server_url": DEFAULT_SERVER_URL,
            "free_points": FREE_POINTS,
        }
    try:
        return json.loads(CONFIG_PATH.read_text("utf-8"))
    except Exception:
        return {
            "server_url": DEFAULT_SERVER_URL,
            "free_points": FREE_POINTS,
        }


def save_config(patch: dict[str, Any]) -> dict[str, Any]:
    with CONFIG_LOCK:
        config = load_config()
        config.update(patch)
        APP_DIR.mkdir(parents=True, exist_ok=True)
        CONFIG_PATH.write_text(json.dumps(config, ensure_ascii=False, indent=2), "utf-8")
        return config


def server_url() -> str:
    return str(load_config().get("server_url") or DEFAULT_SERVER_URL).rstrip("/")


def robust_mac() -> str:
    try:
        result = subprocess.run(
            ["getmac", "/fo", "csv", "/nh"],
            capture_output=True,
            text=True,
            timeout=8,
            creationflags=0x08000000 if os.name == "nt" else 0,
        )
        macs: list[str] = []
        for token in re.findall(r"([0-9A-Fa-f]{2}(?:[-:][0-9A-Fa-f]{2}){5})", result.stdout):
            value = re.sub(r"[^0-9A-Fa-f]", "", token).upper()
            if len(value) == 12 and value != "000000000000":
                macs.append(value)
        if macs:
            return sorted(set(macs))[0]
    except Exception:
        pass
    return f"{uuid.getnode():012X}"


def raw_software_id() -> str:
    return f"{robust_mac()}-{APP_CODE}"


def pretty_serial(raw: str | None = None) -> str:
    value = str(raw or raw_software_id()).upper()
    if value.endswith(f"-{APP_CODE}"):
        value = value[: -len(APP_CODE) - 1]
    hexs = "".join(ch for ch in value if ch.isalnum()).upper()[:12]
    if len(hexs) == 12:
        return "-".join(hexs[i : i + 2] for i in range(0, 12, 2)) + f"-{APP_CODE}"
    return str(raw or raw_software_id())


def pbkdf2_hash(password: str, salt: bytes | None = None) -> dict[str, str]:
    if salt is None:
        salt = os.urandom(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 200_000)
    return {
        "scheme": "pbkdf2_sha256",
        "salt": base64.b64encode(salt).decode("ascii"),
        "hash": base64.b64encode(digest).decode("ascii"),
    }


def verify_password(password: str, stored: dict[str, str] | None) -> bool:
    if not stored:
        return False
    try:
        salt = base64.b64decode(stored["salt"])
        expected = base64.b64decode(stored["hash"])
    except Exception:
        return False
    actual = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 200_000)
    return hmac.compare_digest(actual, expected)


# 服务器是公网直连。绕过系统代理：开 VPN 时系统代理在，关 VPN 后代理不可达会报
# WinError 10061（远程主机拒绝）。强制不走代理，直连服务器，开不开 VPN 都能用。
_DIRECT_OPENER = urllib.request.build_opener(urllib.request.ProxyHandler({}))


def request_json(path: str, body: dict[str, Any] | None = None, token: str = "", timeout: int = 30) -> dict[str, Any]:
    url = server_url() + path
    payload = None if body is None else json.dumps(body).encode("utf-8")
    request = urllib.request.Request(url, data=payload, method="GET" if body is None else "POST")
    request.add_header("Accept", "application/json")
    if body is not None:
        request.add_header("Content-Type", "application/json")
    if token:
        request.add_header("Authorization", f"Bearer {token}")
    with _DIRECT_OPENER.open(request, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def normalize_quota(data: dict[str, Any] | None = None) -> dict[str, int]:
    quota = (data or {}).get("quota") if isinstance(data, dict) else None
    if isinstance(quota, dict):
        return {
            "free": int(quota.get("free") or quota.get("free_pages") or 0),
            "paid": int(quota.get("paid") or quota.get("paid_pages") or 0),
            "available": int(quota.get("available") or quota.get("available_pages") or 0),
            "used": int(quota.get("used") or quota.get("used_pages") or 0),
            "unit": "point",
        }
    config = load_config()
    return {
        "free": int(config.get("free_points") or FREE_POINTS),
        "paid": 0,
        "available": int(config.get("free_points") or FREE_POINTS),
        "used": 0,
        "unit": "point",
    }


def normalize_billing(data: dict[str, Any] | None = None) -> dict[str, Any]:
    billing = (data or {}).get("billing") if isinstance(data, dict) else None
    if isinstance(billing, dict) and billing.get("packages"):
        return billing
    return {
        "contact_wechat": "18033086531",
        "default_points": FREE_POINTS,
        "overdraft_limit": 20,
        "packages": [
            {"points": 50, "standard_price": None, "launch_price": 2.99, "once_per_device": True},
            {"points": 200, "standard_price": 29.9, "launch_price": 9.9, "once_per_device": False},
            {"points": 500, "standard_price": 79.9, "launch_price": 19.9, "once_per_device": False},
            {"points": 1000, "standard_price": 159.9, "launch_price": 29.9, "once_per_device": False},
        ],
        "rules": [
            "JPG、PNG等图片识别：每张1积分",
            "PDF、Word识别：每页1积分",
            "AI智能查找或连续追问：每次成功回答1积分",
            "合同生成：按最终页数每页2积分",
            "任务处理失败、未找到资料或仅进行限制提醒：不扣积分",
        ],
    }


def command_get_serial(_: dict[str, Any]) -> dict[str, Any]:
    return {
        "serial": pretty_serial(),
        "raw": raw_software_id(),
        "app": APP_NAME,
        "unit": "point",
    }


def command_register(_: dict[str, Any]) -> dict[str, Any]:
    serial = raw_software_id()
    try:
        data = request_json(
            "/api/desktop/device/register",
            {
                "software_id": serial,
                "app": APP_NAME,
                "version": APP_VERSION,
                "quota_unit": "point",
                "free_quota": FREE_POINTS,
            },
        )
        STATE["token"] = str(data.get("token") or "")
        return {
            "online": True,
            "serial": pretty_serial(data.get("software_id") or serial),
            "quota": normalize_quota(data),
            "billing": normalize_billing(data),
        }
    except Exception as exc:
        return {
            "online": False,
            "serial": pretty_serial(serial),
            "quota": normalize_quota(),
            "billing": normalize_billing(),
            "message": f"后台暂不可用，已进入本地模式：{str(exc)[:160]}",
        }


def command_sync_quota(_: dict[str, Any]) -> dict[str, Any]:
    if not STATE.get("token"):
        return command_register({})
    try:
        data = request_json("/api/desktop/device/status", token=STATE["token"])
        return {"online": True, "quota": normalize_quota(data), "billing": normalize_billing(data)}
    except Exception as exc:
        return {
            "online": False,
            "quota": normalize_quota(),
            "billing": normalize_billing(),
            "message": f"同步失败，显示本地免费额度：{str(exc)[:160]}",
        }


def command_get_state(_: dict[str, Any]) -> dict[str, Any]:
    config = load_config()
    has_password = bool(config.get("password"))
    return {
        "serial": pretty_serial(),
        "raw_serial": raw_software_id(),
        "has_password": has_password,
        "unlocked": bool(STATE.get("unlocked")) or not has_password,
        "library_dir": config.get("library_dir") or "",
        "server_url": config.get("server_url") or DEFAULT_SERVER_URL,
        "quota": normalize_quota(),
        "billing": normalize_billing(),
    }


def command_set_password(args: dict[str, Any]) -> dict[str, Any]:
    password = str(args.get("password") or "")
    if len(password) < 4:
        raise RuntimeError("本地密码至少 4 位。")
    save_config({"password": pbkdf2_hash(password)})
    STATE["unlocked"] = True
    if str(load_config().get("library_dir") or "").strip():
        schedule_knowledge_index(0)
    return {"has_password": True, "unlocked": True}


def command_verify_password(args: dict[str, Any]) -> dict[str, Any]:
    password = str(args.get("password") or "")
    ok = verify_password(password, load_config().get("password"))
    if ok:
        STATE["unlocked"] = True
        if str(load_config().get("library_dir") or "").strip():
            schedule_knowledge_index(0)
    return {"unlocked": ok}


def require_unlocked() -> None:
    config = load_config()
    if config.get("password") and not STATE.get("unlocked"):
        raise RuntimeError("请先输入本地密码。")


def library_dir() -> Path:
    config = load_config()
    value = str(config.get("library_dir") or "").strip()
    if not value:
        raise RuntimeError("请先设置本地资料库位置。")
    root = Path(value).expanduser().resolve()
    root.mkdir(parents=True, exist_ok=True)
    (root / "files").mkdir(exist_ok=True)
    return root


def db_path(root: Path | None = None) -> Path:
    return (root or library_dir()) / "aidoc.db"


def connect_db(root: Path | None = None) -> sqlite3.Connection:
    # 并发识别只并行读取/调用 AI，写入仍很短；延长 SQLite 等锁时间，避免三个
    # 任务同时完成时因默认 5 秒锁等待而误报失败。
    conn = sqlite3.connect(str(db_path(root)), timeout=30)
    conn.row_factory = sqlite3.Row  # 既能按列名也能按下标取值，兼容旧的 row[0] 写法
    conn.execute("PRAGMA busy_timeout = 30000")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS documents (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          original_path TEXT NOT NULL,
          managed_path TEXT,
          file_name TEXT NOT NULL,
          extension TEXT NOT NULL,
          sha256 TEXT NOT NULL,
          size_bytes INTEGER NOT NULL DEFAULT 0,
          page_count INTEGER NOT NULL DEFAULT 1,
          page_count_method TEXT NOT NULL,
          import_mode TEXT NOT NULL,
          status TEXT NOT NULL DEFAULT 'imported',
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_documents_sha256 ON documents(sha256)")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS assistant_training_notes (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          title TEXT NOT NULL,
          trigger_keywords TEXT NOT NULL DEFAULT '',
          instruction TEXT NOT NULL,
          enabled INTEGER NOT NULL DEFAULT 1,
          use_count INTEGER NOT NULL DEFAULT 0,
          last_used_at TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    conn.commit()
    ensure_recognition_columns(conn)
    ensure_quick_scan_table(conn)
    templatepool.ensure_table(conn)
    knowledge.ensure_tables(conn)
    assistanttasks.ensure_tables(conn)
    return conn


def _knowledge_log(message: str) -> None:
    """内部诊断日志，不在用户界面展示索引实现。"""
    try:
        APP_DIR.mkdir(parents=True, exist_ok=True)
        path = APP_DIR / "knowledge-index.log"
        if path.exists() and path.stat().st_size > 1024 * 1024:
            path.replace(path.with_suffix(".log.1"))
        with path.open("a", encoding="utf-8") as handle:
            handle.write(f"{datetime.now().isoformat(timespec='seconds')} {message}\n")
    except Exception:
        pass


def _knowledge_worker_loop() -> None:
    while True:
        document_id, force = KNOWLEDGE_QUEUE.get()
        try:
            root = library_dir()
            if document_id == 0:
                conn = connect_db(root)
                try:
                    knowledge.cleanup_removed_documents(conn)
                    pending_ids = knowledge.pending_document_ids(conn)
                finally:
                    conn.close()
                _knowledge_log(f"backfill start pending={len(pending_ids)}")
                for pending_id in pending_ids:
                    try:
                        knowledge.index_document(db_path(root), pending_id)
                    except Exception as exc:
                        _knowledge_log(f"document={pending_id} error={type(exc).__name__}:{str(exc)[:240]}")
                    time.sleep(0.02)
                _knowledge_log(f"backfill done indexed={len(pending_ids)}")
            else:
                knowledge.index_document(db_path(root), document_id, force=force)
        except Exception as exc:
            _knowledge_log(f"task={document_id} error={type(exc).__name__}:{str(exc)[:240]}")
        finally:
            with KNOWLEDGE_QUEUE_LOCK:
                KNOWLEDGE_QUEUED.discard((document_id, force))
            KNOWLEDGE_QUEUE.task_done()


def schedule_knowledge_index(document_id: int = 0, *, force: bool = False) -> None:
    global KNOWLEDGE_WORKER
    task = (max(0, int(document_id)), bool(force))
    with KNOWLEDGE_QUEUE_LOCK:
        if task not in KNOWLEDGE_QUEUED:
            KNOWLEDGE_QUEUED.add(task)
            KNOWLEDGE_QUEUE.put(task)
        if KNOWLEDGE_WORKER is None or not KNOWLEDGE_WORKER.is_alive():
            KNOWLEDGE_WORKER = threading.Thread(
                target=_knowledge_worker_loop,
                name="aidoc-knowledge-index",
                daemon=True,
            )
            KNOWLEDGE_WORKER.start()


def command_start_knowledge_index(_: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    root = library_dir()
    conn = connect_db(root)
    try:
        pending = len(knowledge.pending_document_ids(conn))
    finally:
        conn.close()
    schedule_knowledge_index(0)
    return {"started": True, "pending": pending}


def ensure_quick_scan_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS quick_scan_items (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          job_id TEXT NOT NULL,
          source_path TEXT NOT NULL,
          file_name TEXT NOT NULL,
          extension TEXT NOT NULL,
          size_bytes INTEGER NOT NULL DEFAULT 0,
          file_created_at TEXT,
          file_modified_at TEXT,
          page_count INTEGER NOT NULL DEFAULT 1,
          detected_type TEXT NOT NULL DEFAULT 'other',
          detected_type_label TEXT NOT NULL DEFAULT '其他资料',
          first_page_fingerprint TEXT,
          preview_text TEXT,
          status TEXT NOT NULL DEFAULT 'valid',
          reason TEXT,
          duplicate_of_id INTEGER,
          imported_document_id INTEGER,
          scanned_at TEXT NOT NULL
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_quick_scan_job ON quick_scan_items(job_id, id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_quick_scan_status ON quick_scan_items(job_id, status)")
    conn.commit()


def command_set_library_dir(args: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    target = Path(str(args.get("path") or "")).expanduser().resolve()
    if not str(target):
        raise RuntimeError("请选择资料库位置。")
    target.mkdir(parents=True, exist_ok=True)
    (target / "files").mkdir(exist_ok=True)
    connect_db(target).close()
    save_config({"library_dir": str(target)})
    schedule_knowledge_index(0)
    return {"library_dir": str(target), "db": str(db_path(target))}


def command_scan_roots(args: dict[str, Any]) -> dict[str, Any]:
    """返回本机可扫描盘符，供用户一键选择整盘。"""
    roots = []
    if os.name == "nt":
        for code in range(ord("A"), ord("Z") + 1):
            value = f"{chr(code)}:\\"
            if os.path.isdir(value):
                roots.append(value)
    else:
        roots.append("/")
    return {"roots": roots}


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def count_pdf_pages(path: Path) -> tuple[int, str]:
    try:
        from pypdf import PdfReader  # type: ignore

        reader = PdfReader(str(path))
        return max(1, len(reader.pages)), "pypdf"
    except Exception:
        try:
            data = path.read_bytes()
            count = len(re.findall(rb"/Type\s*/Page\b", data))
            return max(1, count), "pdf_regex"
        except Exception:
            return 1, "pdf_fallback"


def count_docx_pages(path: Path) -> tuple[int, str]:
    try:
        from docx import Document  # type: ignore

        document = Document(str(path))
        words = 0
        for paragraph in document.paragraphs:
            words += len(re.findall(r"\S+", paragraph.text))
        return max(1, (words + 899) // 900), "docx_word_estimate"
    except Exception:
        return 1, "docx_fallback"


def count_xlsx_pages(path: Path) -> tuple[int, str]:
    try:
        import openpyxl  # type: ignore

        workbook = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        try:
            return max(1, len(workbook.sheetnames)), "xlsx_sheets"
        finally:
            workbook.close()
    except Exception:
        try:
            with zipfile.ZipFile(path) as archive:
                sheets = [name for name in archive.namelist() if name.startswith("xl/worksheets/sheet")]
                return max(1, len(sheets)), "xlsx_zip_sheets"
        except Exception:
            return 1, "xlsx_fallback"


def count_text_pages(path: Path) -> tuple[int, str]:
    try:
        if path.suffix.lower() == ".csv":
            with path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as handle:
                rows = sum(1 for _ in csv.reader(handle))
            return max(1, (rows + 49) // 50), "csv_50_rows_estimate"
        chars = len(path.read_text("utf-8", errors="ignore"))
        return max(1, (chars + 1799) // 1800), "text_char_estimate"
    except Exception:
        return 1, "text_fallback"


def count_pages(path: Path) -> tuple[int, str]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return count_pdf_pages(path)
    if suffix in IMAGE_EXTENSIONS:
        return 1, "image"
    if suffix == ".docx":
        return count_docx_pages(path)
    if suffix in {".xlsx", ".xlsm"}:
        return count_xlsx_pages(path)
    if suffix in TEXT_EXTENSIONS:
        return count_text_pages(path)
    return 1, "file_fallback"


SCAN_NOISE_DIRS = {
    "$recycle.bin", "system volume information", "windows", "program files",
    "program files (x86)", "programdata", "appdata", "node_modules", ".git", ".svn",
    "__pycache__", ".venv", "venv", "cache", "caches", "temp", "tmp",
}


def iter_supported_files(
    root: Path,
    recursive: bool,
    skip_noise: bool = False,
    exclude_roots: list[Path] | None = None,
) -> list[Path]:
    """用 os.scandir 迭代枚举，扫描整个盘符时明显快于 pathlib.rglob。"""
    files: list[Path] = []
    excluded = [path.resolve() for path in (exclude_roots or [])]
    stack = [root]
    while stack:
        current = stack.pop()
        try:
            resolved_current = current.resolve()
            if any(resolved_current == path or resolved_current.is_relative_to(path) for path in excluded):
                continue
        except OSError:
            continue
        try:
            with os.scandir(current) as entries:
                for entry in entries:
                    try:
                        if entry.is_dir(follow_symlinks=False):
                            if recursive and not (skip_noise and entry.name.casefold() in SCAN_NOISE_DIRS):
                                stack.append(Path(entry.path))
                            continue
                        if not entry.is_file(follow_symlinks=False) or entry.name.startswith("~$"):
                            continue
                        path = Path(entry.path)
                        if path.suffix.lower() in SUPPORTED_EXTENSIONS:
                            files.append(path)
                    except OSError:
                        continue
        except (OSError, PermissionError):
            continue
    files.sort(key=lambda item: str(item).lower())
    return files


def _quick_image_fingerprint(path: Path) -> tuple[str, bool]:
    """返回首页感知指纹和是否疑似空白；相同内容改名后仍能被归为疑似重复。"""
    rendered = docintel.render_pages_png(path, max_pages=1, dpi=110)
    if not rendered:
        return "", True
    try:
        from PIL import Image, ImageOps, ImageStat

        with Image.open(io.BytesIO(rendered[0])) as source:
            gray = ImageOps.grayscale(source)
            sample = gray.copy()
            sample.thumbnail((512, 512))
            stat = ImageStat.Stat(sample)
            pixels = list(getattr(sample, "get_flattened_data", sample.getdata)())
            ink_ratio = sum(1 for value in pixels if value < 245) / max(1, len(pixels))
            blank = float(stat.stddev[0] or 0) < 2.0 or ink_ratio < 0.001

            small = gray.resize((9, 8))
            values = list(getattr(small, "get_flattened_data", small.getdata)())
            bits = []
            for row in range(8):
                offset = row * 9
                bits.extend(values[offset + col] > values[offset + col + 1] for col in range(8))
            number = sum((1 << index) for index, bit in enumerate(bits) if bit)
            return f"image:{number:016x}", blank
    except Exception:
        return "", True


def _quick_scan_fast(path: Path) -> dict[str, Any]:
    stat = path.stat()
    size = int(stat.st_size)
    created = datetime.fromtimestamp(stat.st_ctime).isoformat(timespec="seconds")
    modified = datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds")
    base = {
        "source_path": str(path), "file_name": path.name,
        "extension": path.suffix.lower().lstrip("."), "size_bytes": size,
        "file_created_at": created, "file_modified_at": modified,
        "page_count": 1, "detected_type": "other",
        "detected_type_label": docintel.DOCUMENT_TYPES["other"],
        "first_page_fingerprint": "", "preview_text": "[极速内容指纹]",
        "status": "valid", "reason": "文件结构正常，已建立内容指纹",
    }
    if size <= 0:
        return {**base, "status": "invalid", "reason": "空文件（0 字节）"}

    suffix = path.suffix.lower()
    with path.open("rb") as handle:
        head = handle.read(256 * 1024)
        tail = b""
        if size > 320 * 1024:
            handle.seek(max(0, size - 64 * 1024))
            tail = handle.read(64 * 1024)
    valid_magic = True
    if suffix == ".pdf":
        valid_magic = head.startswith(b"%PDF")
    elif suffix in {".docx", ".xlsx", ".xlsm"}:
        valid_magic = head.startswith(b"PK")
    elif suffix == ".png":
        valid_magic = head.startswith(b"\x89PNG\r\n\x1a\n")
    elif suffix in {".jpg", ".jpeg"}:
        valid_magic = head.startswith(b"\xff\xd8")
    elif suffix == ".bmp":
        valid_magic = head.startswith(b"BM")
    elif suffix in {".tif", ".tiff"}:
        valid_magic = head.startswith((b"II*\x00", b"MM\x00*"))
    elif suffix == ".webp":
        valid_magic = head.startswith(b"RIFF") and head[8:12] == b"WEBP"
    if not valid_magic:
        return {**base, "status": "invalid", "reason": "扩展名与文件内容不匹配或文件已损坏"}

    digest = hashlib.sha256(size.to_bytes(8, "little") + head + tail).hexdigest()
    base["first_page_fingerprint"] = "sample:" + digest
    if suffix in TEXT_EXTENSIONS:
        text = ""
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                text = head.decode(encoding).strip()
                break
            except UnicodeDecodeError:
                continue
        normalized = re.sub(r"\s+", " ", text).strip()
        if not normalized:
            return {**base, "status": "invalid", "reason": "文本文件没有可读取内容"}
        detected = docintel.detect_type(text)
        base.update({
            "detected_type": detected,
            "detected_type_label": docintel.DOCUMENT_TYPES.get(detected, "其他资料"),
            "preview_text": normalized[:240],
        })
    return base


def _quick_scan_one(path: Path) -> dict[str, Any]:
    base = _quick_scan_fast(path)
    if base["status"] != "valid":
        return base
    suffix = path.suffix.lower()
    text = docintel.extract_first_page_text(path)
    normalized = re.sub(r"[^\w\u4e00-\u9fff]+", "", text.lower())
    if len(normalized) >= 8:
        digest = hashlib.sha256(normalized.encode("utf-8")).hexdigest()
        detected = docintel.detect_type(text)
        base.update({
            "detected_type": detected,
            "detected_type_label": docintel.DOCUMENT_TYPES.get(detected, "其他资料"),
            "first_page_fingerprint": "text:" + digest,
            "preview_text": re.sub(r"\s+", " ", text).strip()[:240],
        })
        return base

    if suffix in IMAGE_EXTENSIONS:
        fingerprint, blank = _quick_image_fingerprint(path)
        if fingerprint and not blank:
            return {
                **base, "first_page_fingerprint": fingerprint,
                "preview_text": "[扫描件/图片首页]", "reason": "首页图像可读取，待正式视觉识别",
            }
        return {**base, "status": "invalid", "reason": "首页无法读取或疑似空白"}

    if suffix == ".pdf":
        return {**base, "preview_text": "[扫描版 PDF 首页]", "reason": "扫描版 PDF，已建立内容指纹，待正式视觉识别"}
    return {**base, "reason": "文件结构正常，首页内容待正式识别"}


def _confirm_quick_duplicates(conn: sqlite3.Connection, job_id: str) -> dict[str, int]:
    """只对快速指纹碰撞组读取完整文件；完整 SHA-256 一致才标记为可删除重复件。"""
    rows = conn.execute(
        """
        SELECT * FROM quick_scan_items
        WHERE job_id = ? AND first_page_fingerprint IN (
          SELECT first_page_fingerprint FROM quick_scan_items
          WHERE job_id = ? AND first_page_fingerprint <> ''
          GROUP BY first_page_fingerprint HAVING COUNT(*) > 1
        )
        ORDER BY id
        """,
        (job_id, job_id),
    ).fetchall()
    if rows:
        paths = {int(row["id"]): Path(str(row["source_path"])) for row in rows}
        digests: dict[int, str] = {}
        errors: dict[int, str] = {}

        def calculate(item: tuple[int, Path]) -> tuple[int, str, str]:
            item_id, path = item
            try:
                return item_id, file_sha256(path), ""
            except Exception as exc:
                return item_id, "", str(exc)[:200]

        with ThreadPoolExecutor(max_workers=min(4, len(paths)), thread_name_prefix="aidoc-hash") as executor:
            for item_id, digest, error in executor.map(calculate, paths.items()):
                if digest:
                    digests[item_id] = digest
                else:
                    errors[item_id] = error or "无法读取完整文件"

        groups: dict[str, list[sqlite3.Row]] = {}
        for row in rows:
            groups.setdefault(str(row["first_page_fingerprint"]), []).append(row)
        for group in groups.values():
            hash_heads: dict[str, int] = {}
            first_id = int(group[0]["id"])
            for row in group:
                item_id = int(row["id"])
                if item_id in errors:
                    conn.execute(
                        "UPDATE quick_scan_items SET status='error', reason=?, duplicate_of_id=NULL WHERE id=?",
                        ("完整文件校验失败：" + errors[item_id], item_id),
                    )
                    continue
                digest = digests[item_id]
                if digest in hash_heads:
                    conn.execute(
                        """UPDATE quick_scan_items
                           SET status='duplicate_exact', reason='完整 SHA-256 一致，确认是完全重复文件', duplicate_of_id=?
                           WHERE id=?""",
                        (hash_heads[digest], item_id),
                    )
                else:
                    hash_heads[digest] = item_id
                    if item_id == first_id:
                        conn.execute(
                            "UPDATE quick_scan_items SET status='valid', duplicate_of_id=NULL WHERE id=?", (item_id,)
                        )
                    else:
                        conn.execute(
                            """UPDATE quick_scan_items
                               SET status='similar', reason='快速内容指纹相似，但完整哈希不同，请人工复核', duplicate_of_id=NULL
                               WHERE id=?""",
                            (item_id,),
                        )
        conn.commit()

    stat_rows = conn.execute(
        "SELECT status, COUNT(*) AS amount FROM quick_scan_items WHERE job_id=? GROUP BY status", (job_id,)
    ).fetchall()
    return {str(row["status"]): int(row["amount"]) for row in stat_rows}


@dataclass
class ImportResult:
    created: bool
    document_id: int
    managed_path: str


def import_document(conn: sqlite3.Connection, src: Path, root: Path, digest: str, pages: int, method: str, import_mode: str) -> ImportResult:
    now = datetime.now().isoformat(timespec="seconds")
    existing = conn.execute("SELECT id, managed_path FROM documents WHERE sha256 = ?", (digest,)).fetchone()
    if existing:
        return ImportResult(False, int(existing[0]), str(existing[1] or ""))

    managed_path = ""
    if import_mode == "copy":
        stamp = datetime.now().strftime("%Y/%m")
        target_dir = root / "files" / stamp
        target_dir.mkdir(parents=True, exist_ok=True)
        target = target_dir / f"{digest[:16]}{src.suffix.lower()}"
        if not target.exists():
            shutil.copy2(src, target)
        managed_path = str(target)

    cursor = conn.execute(
        """
        INSERT INTO documents (
          original_path, managed_path, file_name, extension, sha256, size_bytes,
          page_count, page_count_method, import_mode, status, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'imported', ?, ?)
        """,
        (
            str(src),
            managed_path,
            src.name,
            src.suffix.lower().lstrip("."),
            digest,
            src.stat().st_size,
            pages,
            method,
            import_mode,
            now,
            now,
        ),
    )
    conn.commit()
    return ImportResult(True, int(cursor.lastrowid), managed_path)


def command_scan_folder(args: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    source = Path(str(args.get("path") or "")).expanduser().resolve()
    if not source.exists() or not source.is_dir():
        raise RuntimeError("请选择有效的文件夹。")
    recursive = bool(args.get("recursive", True))
    import_mode = str(args.get("import_mode") or "copy")
    if import_mode not in {"copy", "index"}:
        import_mode = "copy"

    root = library_dir()
    conn = connect_db(root)
    job_id = uuid.uuid4().hex
    started = time.time()

    emit({
        "event": "aidoc_progress",
        "job_id": job_id,
        "stage": "scan",
        "stage_label": "正在扫描文件夹",
        "done_pages": 0,
        "total_pages": 0,
        "current_file": "",
    })

    files = iter_supported_files(source, recursive)
    total_files = len(files)
    total_pages = 0
    success_pages = 0
    skipped_pages = 0
    failed_pages = 0
    created_documents = 0
    skipped_documents = 0
    rows: list[dict[str, Any]] = []

    for index, file_path in enumerate(files, 1):
        rel = str(file_path.relative_to(source)) if file_path.is_relative_to(source) else file_path.name
        emit({
            "event": "aidoc_progress",
            "job_id": job_id,
            "stage": "count",
            "stage_label": "正在统计页数",
            "file_index": index,
            "file_total": total_files,
            "current_file": rel,
            "done_pages": total_pages,
            "total_pages": None,
        })
        try:
            pages, method = count_pages(file_path)
            digest = file_sha256(file_path)
            total_pages += pages
            result = import_document(conn, file_path, root, digest, pages, method, import_mode)
            if result.created:
                created_documents += 1
                success_pages += pages
                status = "imported"
            else:
                skipped_documents += 1
                skipped_pages += pages
                status = "duplicate"
            rows.append({
                "id": result.document_id,
                "name": file_path.name,
                "path": str(file_path),
                "managed_path": result.managed_path,
                "extension": file_path.suffix.lower().lstrip("."),
                "pages": pages,
                "method": method,
                "status": status,
            })
            emit({
                "event": "aidoc_progress",
                "job_id": job_id,
                "stage": "import",
                "stage_label": "正在导入资料库",
                "file_index": index,
                "file_total": total_files,
                "current_file": rel,
                "current_file_pages": pages,
                "done_pages": success_pages + skipped_pages + failed_pages,
                "total_pages": total_pages,
                "success_pages": success_pages,
                "skipped_pages": skipped_pages,
                "failed_pages": failed_pages,
            })
        except Exception as exc:
            failed_pages += 1
            rows.append({
                "id": None,
                "name": file_path.name,
                "path": str(file_path),
                "extension": file_path.suffix.lower().lstrip("."),
                "pages": 1,
                "method": "failed",
                "status": "failed",
                "error": str(exc)[:240],
            })
            emit({
                "event": "aidoc_progress",
                "job_id": job_id,
                "stage": "failed",
                "stage_label": "文件处理失败",
                "file_index": index,
                "file_total": total_files,
                "current_file": rel,
                "done_pages": success_pages + skipped_pages + failed_pages,
                "total_pages": total_pages,
                "success_pages": success_pages,
                "skipped_pages": skipped_pages,
                "failed_pages": failed_pages,
                "error": str(exc)[:240],
            })

    conn.close()
    elapsed = max(0.001, time.time() - started)
    emit({
        "event": "aidoc_progress",
        "job_id": job_id,
        "stage": "done",
        "stage_label": "已完成",
        "done_pages": success_pages + skipped_pages + failed_pages,
        "total_pages": total_pages,
        "success_pages": success_pages,
        "skipped_pages": skipped_pages,
        "failed_pages": failed_pages,
    })

    return {
        "job_id": job_id,
        "source": str(source),
        "library_dir": str(root),
        "file_count": total_files,
        "created_documents": created_documents,
        "skipped_documents": skipped_documents,
        "total_pages": total_pages,
        "billable_pages": success_pages,
        "skipped_pages": skipped_pages,
        "failed_pages": failed_pages,
        "elapsed_seconds": round(elapsed, 2),
        "items": rows[-200:],
    }


def _date_boundary(value: Any, end_of_day: bool = False) -> float | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        parsed = datetime.fromisoformat(text)
        if end_of_day and len(text) <= 10:
            parsed = parsed.replace(hour=23, minute=59, second=59)
        return parsed.timestamp()
    except ValueError:
        return None


def command_quick_scan(args: dict[str, Any]) -> dict[str, Any]:
    """只读每份文件的首页做本地预检，不调用 AI、不扣积分、不自动删除。"""
    require_unlocked()
    source = Path(str(args.get("path") or "")).expanduser().resolve()
    if not source.is_dir():
        raise RuntimeError("请选择有效的扫描文件夹。")
    recursive = bool(args.get("recursive", True))
    scan_depth = str(args.get("scan_depth") or "fast")
    if scan_depth not in {"fast", "preview"}:
        scan_depth = "fast"
    skip_noise = bool(args.get("skip_noise", True))
    selected_exts = {
        (str(value).lower() if str(value).startswith(".") else "." + str(value).lower())
        for value in list(args.get("extensions") or [])
    }
    selected_exts &= SUPPORTED_EXTENSIONS
    if not selected_exts:
        selected_exts = set(SUPPORTED_EXTENSIONS)
    try:
        min_size = max(0, int(float(args.get("min_size_mb") or 0) * 1024 * 1024))
    except (TypeError, ValueError):
        min_size = 0
    try:
        max_size_value = float(args.get("max_size_mb") or 0)
        max_size = int(max_size_value * 1024 * 1024) if max_size_value > 0 else None
    except (TypeError, ValueError):
        max_size = None
    created_from = _date_boundary(args.get("created_from"))
    created_to = _date_boundary(args.get("created_to"), end_of_day=True)

    library_root = library_dir()
    all_files = iter_supported_files(source, recursive, skip_noise=skip_noise, exclude_roots=[library_root])
    files: list[Path] = []
    filtered_out = 0
    for path in all_files:
        try:
            stat = path.stat()
        except OSError:
            filtered_out += 1
            continue
        if path.suffix.lower() not in selected_exts:
            filtered_out += 1
            continue
        if stat.st_size < min_size or (max_size is not None and stat.st_size > max_size):
            filtered_out += 1
            continue
        if created_from is not None and stat.st_ctime < created_from:
            filtered_out += 1
            continue
        if created_to is not None and stat.st_ctime > created_to:
            filtered_out += 1
            continue
        files.append(path)

    job_id = uuid.uuid4().hex
    conn = connect_db(library_root)
    old_jobs = conn.execute(
        "SELECT job_id FROM quick_scan_items GROUP BY job_id ORDER BY MAX(id) DESC LIMIT -1 OFFSET 10"
    ).fetchall()
    for old_job in old_jobs:
        conn.execute("DELETE FROM quick_scan_items WHERE job_id = ?", (old_job[0],))
    if old_jobs:
        conn.commit()
    scanned_at = datetime.now().isoformat(timespec="seconds")
    fingerprints: dict[str, tuple[int, str]] = {}
    counts = {"valid": 0, "duplicate": 0, "invalid": 0, "error": 0}
    completed = 0
    workers = min(10 if scan_depth == "fast" else 2, max(1, len(files)))
    worker_fn = _quick_scan_fast if scan_depth == "fast" else _quick_scan_one
    emit({
        "event": "aidoc_quick_scan", "job_id": job_id, "stage": "start",
        "stage_label": (
            f"正在极速扫描整个位置（{workers} 路内容指纹）…" if scan_depth == "fast"
            else f"正在首页预检（{workers} 路）…"
        ),
        "done": 0, "total": len(files), "filtered_out": filtered_out,
    })
    try:
        with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="aidoc-quick") as executor:
            # 工作仍并发执行，但按已排序路径落库，确保每组重复文件的“保留件”稳定可预测。
            futures = [(path, executor.submit(worker_fn, path)) for path in files]
            for path, future in futures:
                try:
                    item = future.result()
                except Exception as exc:
                    stat = path.stat()
                    item = {
                        "source_path": str(path), "file_name": path.name,
                        "extension": path.suffix.lower().lstrip("."),
                        "size_bytes": int(stat.st_size),
                        "file_created_at": datetime.fromtimestamp(stat.st_ctime).isoformat(timespec="seconds"),
                        "file_modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
                        "page_count": 1, "detected_type": "other", "detected_type_label": "其他资料",
                        "first_page_fingerprint": "", "preview_text": "",
                        "status": "error", "reason": str(exc)[:240],
                    }
                fingerprint = str(item.get("first_page_fingerprint") or "")
                duplicate_of_id = None
                if fingerprint and item["status"] == "valid" and fingerprint in fingerprints:
                    duplicate_of_id, duplicate_name = fingerprints[fingerprint]
                    item["status"] = "duplicate"
                    item["reason"] = (
                        f"文件内容指纹与《{duplicate_name}》相同，疑似重复" if scan_depth == "fast"
                        else f"首页内容与《{duplicate_name}》相同，疑似重复"
                    )

                cursor = conn.execute(
                    """
                    INSERT INTO quick_scan_items (
                      job_id, source_path, file_name, extension, size_bytes,
                      file_created_at, file_modified_at, page_count,
                      detected_type, detected_type_label, first_page_fingerprint,
                      preview_text, status, reason, duplicate_of_id, scanned_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        job_id, item["source_path"], item["file_name"], item["extension"], item["size_bytes"],
                        item["file_created_at"], item["file_modified_at"], item["page_count"],
                        item["detected_type"], item["detected_type_label"], fingerprint,
                        item["preview_text"], item["status"], item["reason"], duplicate_of_id, scanned_at,
                    ),
                )
                item_id = int(cursor.lastrowid)
                if fingerprint and fingerprint not in fingerprints and item["status"] == "valid":
                    fingerprints[fingerprint] = (item_id, item["file_name"])
                counts[item["status"]] += 1
                completed += 1
                if completed % 50 == 0:
                    conn.commit()
                emit({
                    "event": "aidoc_quick_scan", "job_id": job_id, "stage": "scan",
                    "stage_label": "正在建立文件内容指纹…" if scan_depth == "fast" else "正在读取每份文档的首页…", "done": completed,
                    "total": len(files), "current_file": item["file_name"], **counts,
                })
        conn.commit()
        emit({
            "event": "aidoc_quick_scan", "job_id": job_id, "stage": "verify_duplicates",
            "stage_label": "正在用完整 SHA-256 确认重复文件…", "done": completed,
            "total": len(files), **counts,
        })
        confirmed = _confirm_quick_duplicates(conn, job_id)
        counts = {
            "valid": int(confirmed.get("valid", 0)),
            "duplicate_exact": int(confirmed.get("duplicate_exact", 0)),
            "similar": int(confirmed.get("similar", 0)),
            "invalid": int(confirmed.get("invalid", 0)),
            "error": int(confirmed.get("error", 0)),
        }
        # duplicate 保留为兼容字段，含义已收紧为“完整哈希确认重复”。
        counts["duplicate"] = counts["duplicate_exact"]
    finally:
        conn.close()

    emit({
        "event": "aidoc_quick_scan", "job_id": job_id, "stage": "done",
        "stage_label": "快速预检完成，请审核结果", "done": completed,
        "total": len(files), "filtered_out": filtered_out, **counts,
    })
    return {
        "job_id": job_id, "source": str(source), "total": len(files), "scan_depth": scan_depth,
        "filtered_out": filtered_out, "workers": workers, **counts,
    }


def command_quick_scan_list(args: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    job_id = str(args.get("job_id") or "").strip()
    if not job_id:
        raise RuntimeError("缺少快速扫描任务编号。")
    status = str(args.get("status") or "").strip()
    page = max(1, int(args.get("page") or 1))
    per_page = max(10, min(200, int(args.get("per_page") or 50)))
    where = ["job_id = ?"]
    params: list[Any] = [job_id]
    if status:
        where.append("status = ?")
        params.append(status)
    conn = connect_db(library_dir())
    clause = " AND ".join(where)
    total = int(conn.execute(f"SELECT COUNT(*) FROM quick_scan_items WHERE {clause}", params).fetchone()[0])
    page = min(page, max(1, (total + per_page - 1) // per_page))
    rows = conn.execute(
        f"SELECT * FROM quick_scan_items WHERE {clause} ORDER BY id LIMIT ? OFFSET ?",
        [*params, per_page, (page - 1) * per_page],
    ).fetchall()
    stat_rows = conn.execute(
        "SELECT status, COUNT(*) AS amount FROM quick_scan_items WHERE job_id = ? GROUP BY status", (job_id,)
    ).fetchall()
    conn.close()
    return {
        "items": [dict(row) for row in rows], "total": total, "page": page,
        "per_page": per_page, "stats": {str(row["status"]): int(row["amount"]) for row in stat_rows},
    }


def command_quick_scan_mark(args: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    status = str(args.get("status") or "")
    if status not in {"valid", "invalid"}:
        raise RuntimeError("只能标记为有效或无效。")
    ids = list(dict.fromkeys(int(value) for value in list(args.get("ids") or []) if str(value).isdigit()))[:1000]
    if not ids:
        return {"updated": 0}
    conn = connect_db(library_dir())
    placeholders = ",".join("?" for _ in ids)
    reason = "用户审核后标记有效" if status == "valid" else "用户审核后标记无效"
    cursor = conn.execute(
        f"UPDATE quick_scan_items SET status = ?, reason = ? WHERE id IN ({placeholders}) AND status <> 'deleted'",
        [status, reason, *ids],
    )
    conn.commit()
    conn.close()
    return {"updated": int(cursor.rowcount)}


def command_quick_scan_delete(args: dict[str, Any]) -> dict[str, Any]:
    """只删除无效或完整 SHA-256 已确认重复的源文件；相似件必须人工标为无效。"""
    require_unlocked()
    ids = list(dict.fromkeys(int(value) for value in list(args.get("ids") or []) if str(value).isdigit()))[:1000]
    job_id = str(args.get("job_id") or "").strip()
    all_discard = bool(args.get("all_discard")) and bool(job_id)
    if not ids and not all_discard:
        return {"deleted": 0, "failed": []}
    conn = connect_db(library_dir())
    if all_discard:
        rows = conn.execute(
            "SELECT * FROM quick_scan_items WHERE job_id = ? AND status IN ('invalid','duplicate_exact')", (job_id,)
        ).fetchall()
    else:
        placeholders = ",".join("?" for _ in ids)
        rows = conn.execute(
            f"SELECT * FROM quick_scan_items WHERE id IN ({placeholders}) AND status IN ('invalid','duplicate_exact')", ids
        ).fetchall()
    deleted = 0
    failed: list[dict[str, Any]] = []
    for row in rows:
        path = Path(str(row["source_path"] or ""))
        try:
            if path.exists():
                if not path.is_file():
                    raise RuntimeError("目标不是文件")
                path.unlink()
            conn.execute("UPDATE quick_scan_items SET status='deleted', reason='用户审核后已删除' WHERE id=?", (row["id"],))
            deleted += 1
        except Exception as exc:
            failed.append({"id": int(row["id"]), "file_name": str(row["file_name"]), "error": str(exc)[:200]})
    conn.commit()
    conn.close()
    return {"deleted": deleted, "failed": failed}


def command_quick_scan_import(args: dict[str, Any]) -> dict[str, Any]:
    """审核后把有效项目正式导入资料库；此时才读取完整文件做精确 hash。"""
    require_unlocked()
    ids = list(dict.fromkeys(int(value) for value in list(args.get("ids") or []) if str(value).isdigit()))[:1000]
    job_id = str(args.get("job_id") or "").strip()
    all_valid = bool(args.get("all_valid")) and bool(job_id)
    import_mode = str(args.get("import_mode") or "copy")
    if import_mode not in {"copy", "index"}:
        import_mode = "copy"
    if not ids and not all_valid:
        return {"created": 0, "duplicates": 0, "document_ids": [], "failed": []}
    root = library_dir()
    conn = connect_db(root)
    if all_valid:
        rows = conn.execute(
            "SELECT * FROM quick_scan_items WHERE job_id = ? AND status = 'valid' ORDER BY id", (job_id,)
        ).fetchall()
    else:
        placeholders = ",".join("?" for _ in ids)
        rows = conn.execute(
            f"SELECT * FROM quick_scan_items WHERE id IN ({placeholders}) AND status = 'valid' ORDER BY id", ids
        ).fetchall()
    created = duplicates = 0
    document_ids: list[int] = []
    failed: list[dict[str, Any]] = []
    for row in rows:
        path = Path(str(row["source_path"] or ""))
        try:
            if not path.is_file():
                raise RuntimeError("源文件已不存在")
            pages, method = count_pages(path)
            digest = file_sha256(path)
            result = import_document(conn, path, root, digest, pages, method, import_mode)
            created += int(result.created)
            duplicates += int(not result.created)
            document_ids.append(result.document_id)
            conn.execute(
                "UPDATE quick_scan_items SET status='imported', imported_document_id=?, reason=? WHERE id=?",
                (result.document_id, "已正式导入" if result.created else "完整文件重复，已关联库内资料", row["id"]),
            )
        except Exception as exc:
            failed.append({"id": int(row["id"]), "file_name": str(row["file_name"]), "error": str(exc)[:200]})
    conn.commit()
    conn.close()
    return {
        "created": created, "duplicates": duplicates,
        "document_ids": list(dict.fromkeys(document_ids)), "failed": failed,
    }


def command_import_files(args: dict[str, Any]) -> dict[str, Any]:
    """把拖入/选中的具体文件导入资料库（不是扫文件夹）。args: {paths:[...], import_mode}
    默认 copy 模式：拷一份进库，原文件不动。复用 import_document + 同一套进度事件。"""
    require_unlocked()
    raw = args.get("paths") or []
    import_mode = str(args.get("import_mode") or "copy")
    if import_mode not in {"copy", "index"}:
        import_mode = "copy"

    # 过滤成支持的文件类型；去重
    files: list[Path] = []
    seen: set[str] = set()
    for item in raw:
        p = Path(str(item)).expanduser()
        key = str(p).lower()
        if key in seen:
            continue
        seen.add(key)
        if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS:
            files.append(p)
    if not files:
        raise RuntimeError("没有可导入的文件（支持 PDF / 图片 / Word / Excel）。")

    root = library_dir()
    conn = connect_db(root)
    job_id = uuid.uuid4().hex
    started = time.time()
    total_files = len(files)
    total_pages = success_pages = skipped_pages = failed_pages = 0
    created_documents = skipped_documents = 0
    rows: list[dict[str, Any]] = []

    for index, file_path in enumerate(files, 1):
        emit({"event": "aidoc_progress", "job_id": job_id, "stage": "import",
              "stage_label": "正在导入资料库", "file_index": index, "file_total": total_files,
              "current_file": file_path.name, "done_pages": total_pages, "total_pages": None})
        try:
            pages, method = count_pages(file_path)
            digest = file_sha256(file_path)
            total_pages += pages
            result = import_document(conn, file_path, root, digest, pages, method, import_mode)
            if result.created:
                created_documents += 1
                success_pages += pages
                status = "imported"
            else:
                skipped_documents += 1
                skipped_pages += pages
                status = "duplicate"
            rows.append({"id": result.document_id, "name": file_path.name, "path": str(file_path),
                         "managed_path": result.managed_path,
                         "extension": file_path.suffix.lower().lstrip("."),
                         "pages": pages, "method": method, "status": status})
        except Exception as exc:
            failed_pages += 1
            rows.append({"id": None, "name": file_path.name, "path": str(file_path),
                         "extension": file_path.suffix.lower().lstrip("."),
                         "pages": 1, "method": "failed", "status": "failed", "error": str(exc)[:240]})
        emit({"event": "aidoc_progress", "job_id": job_id, "stage": "import",
              "stage_label": "正在导入资料库", "file_index": index, "file_total": total_files,
              "current_file": file_path.name,
              "done_pages": success_pages + skipped_pages + failed_pages, "total_pages": total_pages,
              "success_pages": success_pages, "skipped_pages": skipped_pages, "failed_pages": failed_pages})

    conn.close()
    emit({"event": "aidoc_progress", "job_id": job_id, "stage": "done", "stage_label": "已完成",
          "done_pages": success_pages + skipped_pages + failed_pages, "total_pages": total_pages,
          "success_pages": success_pages, "skipped_pages": skipped_pages, "failed_pages": failed_pages})

    return {
        "job_id": job_id, "library_dir": str(root), "file_count": total_files,
        "created_documents": created_documents, "skipped_documents": skipped_documents,
        "total_pages": total_pages, "billable_pages": success_pages,
        "skipped_pages": skipped_pages, "failed_pages": failed_pages,
        "elapsed_seconds": round(max(0.001, time.time() - started), 2), "items": rows[-200:],
    }


RECOGNITION_COLUMNS = [
    ("document_type", "TEXT"),
    ("document_type_label", "TEXT"),
    ("company_name", "TEXT"),
    ("brand", "TEXT"),
    ("certificate_no", "TEXT"),
    ("issuer", "TEXT"),
    ("issued_at", "TEXT"),
    ("expires_at", "TEXT"),
    ("applicable_scope", "TEXT"),
    ("extra_fields", "TEXT"),
    ("tags", "TEXT"),
    ("ai_summary", "TEXT"),
    ("ai_confidence", "INTEGER"),
    ("review_status", "TEXT"),
    ("duplicate_of_id", "INTEGER"),
    ("duplicate_reason", "TEXT"),
    ("recognized_at", "TEXT"),
]


def ensure_recognition_columns(conn: sqlite3.Connection) -> None:
    """给老库补识别相关列（sqlite 没有 ADD COLUMN IF NOT EXISTS，按 pragma 判断）。"""
    have = {row[1] for row in conn.execute("PRAGMA table_info(documents)").fetchall()}
    for name, col_type in RECOGNITION_COLUMNS:
        if name not in have:
            conn.execute(f"ALTER TABLE documents ADD COLUMN {name} {col_type}")
    conn.commit()


def _set_recognize_error(message: str) -> None:
    """保存当前识别线程的错误，避免并发任务互相覆盖失败原因。"""
    RECOGNIZE_CONTEXT.last_error = message


def _get_recognize_error() -> str:
    return str(getattr(RECOGNIZE_CONTEXT, "last_error", "") or "")


def recognize_via_server(mode: str, text: str, images: list[bytes], page_count: int, filename: str) -> dict[str, Any] | None:
    """把证据（文字或页图 base64）发服务器做 AI 识别，按页计费。
    失败返回 None 并保存当前线程的真实原因，调用方退回纯规则识别。"""
    _set_recognize_error("")
    token = STATE.get("token") or ""
    if not token:
        _set_recognize_error("未注册或离线（没有 token）")
        return None
    body: dict[str, Any] = {
        "mode": mode,
        "filename": filename,
        "page_count": int(page_count),
        "instruction": docintel.extraction_instruction(),
    }
    if mode == "vision":
        body["images"] = [base64.b64encode(b).decode("ascii") for b in images[:4]]
    else:
        body["text"] = text[:20000]
    try:
        resp = request_json("/api/desktop/document/recognize", body, token=token, timeout=300)
    except urllib.error.HTTPError as exc:
        try:
            detail = exc.read().decode("utf-8", "replace")
            msg = json.loads(detail).get("message") or detail
        except Exception:
            msg = ""
        _set_recognize_error(f"服务器 HTTP {exc.code}：{str(msg)[:240]}")
        return None
    except Exception as exc:
        _set_recognize_error(f"{type(exc).__name__}：{str(exc)[:240]}")
        return None
    if not isinstance(resp, dict):
        _set_recognize_error("服务器返回不是 JSON 对象")
        return None
    data = resp.get("data") if isinstance(resp.get("data"), dict) else resp
    content = data.get("content") if isinstance(data, dict) else None
    if isinstance(content, str):
        parsed = docintel.parse_ai_json(content)
        if parsed is None:
            _set_recognize_error("AI 返回的不是有效 JSON：" + content[:160])
        return parsed
    return data if isinstance(data, dict) else None


def _doc_row_for_analyze(conn: sqlite3.Connection, doc_id: int):
    return conn.execute(
        "SELECT id, original_path, managed_path, file_name, extension, page_count FROM documents WHERE id = ?",
        (doc_id,),
    ).fetchone()


def _mark_duplicate_by_cert(conn: sqlite3.Connection, doc_id: int, cert_no: str) -> int | None:
    """同证件号查重：命中更早的同号资料则标记为重复（checksum 重复在扫描导入时已挡）。"""
    cert_no = (cert_no or "").strip()
    if len(cert_no) < 4:
        return None
    row = conn.execute(
        "SELECT id FROM documents WHERE certificate_no = ? AND id <> ? ORDER BY id ASC LIMIT 1",
        (cert_no, doc_id),
    ).fetchone()
    if row:
        conn.execute(
            "UPDATE documents SET duplicate_of_id = ?, duplicate_reason = 'certificate' WHERE id = ?",
            (int(row[0]), doc_id),
        )
        return int(row[0])
    return None


def command_analyze_document(args: dict[str, Any]) -> dict[str, Any]:
    """对一份已导入资料做 AI 识别归类，结果进「待确认」。
    args: {id} 资料库里的文档 id。"""
    require_unlocked()
    doc_id = int(args.get("id") or 0)
    if doc_id <= 0:
        raise RuntimeError("缺少资料 id。")

    root = library_dir()
    conn = connect_db(root)
    ensure_recognition_columns(conn)
    row = _doc_row_for_analyze(conn, doc_id)
    if not row:
        conn.close()
        raise RuntimeError("资料不存在。")

    managed, original = str(row[2] or ""), str(row[1] or "")
    path = Path(managed) if managed and Path(managed).exists() else Path(original)
    if not path.exists():
        conn.close()
        raise RuntimeError("源文件已不在原位置，无法识别。")

    def _emit(stage: str, label: str) -> None:
        emit({"event": "aidoc_recognize", "id": doc_id, "file": path.name,
              "stage": stage, "stage_label": label})

    _emit("read", "正在读取文件内容…")
    page_count = int(row[5] or 1)
    text = docintel.extract_text(path)
    use_vision = docintel.needs_vision(path, text)
    rule = docintel.rule_suggestion(text)

    ai = None
    source = "rule"
    ai_error = ""
    if use_vision:
        _emit("render", "正在转换页面图像…")
        images = docintel.render_pages_png(path, max_pages=min(2, page_count))
        if images:
            _emit("ai", "AI 正在看图识别（约 10 秒）…")
            ai = recognize_via_server("vision", "", images, page_count, path.name)
            if ai is not None:
                source = "vision_ai"
            else:
                ai_error = _get_recognize_error() or "AI 识别未返回结果"
        else:
            ai_error = "无法把文件转成可识别的页面图像"
    elif len(text) >= 30:
        _emit("ai", "AI 正在阅读文字识别…")
        ai = recognize_via_server("text", text, [], page_count, path.name)
        if ai is not None:
            source = "text_ai"
        else:
            ai_error = _get_recognize_error() or "AI 识别未返回结果"
    else:
        ai_error = "文件里几乎没有可识别的文字，也无法看图（已用规则粗分类）"

    _emit("save", "正在保存识别结果…")

    suggestion = docintel.merge_rule_and_ai(rule, ai)
    now = datetime.now().isoformat(timespec="seconds")
    conn.execute(
        """
        UPDATE documents SET
          document_type = ?, document_type_label = ?, company_name = ?, brand = ?,
          certificate_no = ?, issuer = ?, issued_at = ?, expires_at = ?, applicable_scope = ?,
          extra_fields = ?, tags = ?, ai_summary = ?, ai_confidence = ?,
          review_status = 'pending_review', recognized_at = ?, updated_at = ?
        WHERE id = ?
        """,
        (
            suggestion["document_type"], suggestion["document_type_label"],
            suggestion["company_name"], suggestion["brand"], suggestion["certificate_no"],
            suggestion["issuer"], suggestion["issued_at"], suggestion["expires_at"],
            suggestion["applicable_scope"], json.dumps(suggestion["extra_fields"], ensure_ascii=False),
            json.dumps(suggestion["tags"], ensure_ascii=False), suggestion["ai_summary"],
            suggestion["ai_confidence"], now, now, doc_id,
        ),
    )
    conn.commit()
    duplicate_of = _mark_duplicate_by_cert(conn, doc_id, suggestion["certificate_no"])
    template_created = None
    try:
        template_created = templatepool.upsert_from_document(conn, doc_id, active=False)
    except Exception:
        # 模板提炼失败不影响资料识别；用户仍可在资源池中手动重试。
        template_created = None
    conn.commit()
    conn.close()
    schedule_knowledge_index(doc_id)
    _emit("done", "识别完成")

    return {
        "id": doc_id,
        "source": source,
        "used_vision": use_vision,
        "ai_available": ai is not None,
        "ai_error": ai_error,
        "suggestion": suggestion,
        "duplicate_of_id": duplicate_of,
        "review_status": "pending_review",
        "template_candidate": template_created,
    }


def command_analyze_documents(args: dict[str, Any]) -> dict[str, Any]:
    """有限并发识别多份资料；单份识别的模型、提示词和证据处理完全不变。"""
    require_unlocked()
    ids: list[int] = []
    seen: set[int] = set()
    for value in list(args.get("ids") or [])[:5000]:
        try:
            doc_id = int(value)
        except (TypeError, ValueError):
            continue
        if doc_id > 0 and doc_id not in seen:
            seen.add(doc_id)
            ids.append(doc_id)
    if not ids:
        raise RuntimeError("请选择需要识别的资料。")

    try:
        requested_workers = int(args.get("workers") or 3)
    except (TypeError, ValueError):
        requested_workers = 3
    workers = max(1, min(4, requested_workers, len(ids)))
    # 先完成旧资料库的建表/补列，再登记可取消任务，避免初始化失败留下僵尸任务。
    connect_db(library_dir()).close()
    job_id = re.sub(r"[^a-zA-Z0-9_-]", "", str(args.get("job_id") or ""))[:80] or uuid.uuid4().hex
    cancel_event = threading.Event()
    job_state: dict[str, Any] = {"cancel": cancel_event, "futures": []}
    with RECOGNITION_JOBS_LOCK:
        if job_id in RECOGNITION_JOBS:
            raise RuntimeError("识别任务编号重复，请重新开始。")
        RECOGNITION_JOBS[job_id] = job_state
    total = len(ids)
    completed = 0
    ai_ok = 0
    ai_fail = 0
    cancelled = 0
    items_by_id: dict[int, dict[str, Any]] = {}

    try:
        emit({
            "event": "aidoc_recognize_batch", "job_id": job_id,
            "stage": "start", "stage_label": f"正在并发识别（{workers} 路）…",
            "done": 0, "total": total, "workers": workers,
        })
        with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="aidoc-recognize") as executor:
            futures = {executor.submit(command_analyze_document, {"id": doc_id}): doc_id for doc_id in ids}
            job_state["futures"] = list(futures)
            for future in as_completed(futures):
                doc_id = futures[future]
                try:
                    result = future.result()
                    if bool(result.get("ai_available")):
                        ai_ok += 1
                    else:
                        ai_fail += 1
                    items_by_id[doc_id] = {"id": doc_id, "ok": True, "data": result}
                    error = str(result.get("ai_error") or "")
                except CancelledError:
                    cancelled += 1
                    error = "用户已取消"
                    items_by_id[doc_id] = {"id": doc_id, "ok": False, "cancelled": True, "error": error}
                except Exception as exc:
                    ai_fail += 1
                    error = str(exc)[:500]
                    items_by_id[doc_id] = {"id": doc_id, "ok": False, "error": error}
                completed += 1
                stopping = cancel_event.is_set()
                emit({
                    "event": "aidoc_recognize_batch", "job_id": job_id,
                    "stage": "cancelling" if stopping else "progress",
                    "stage_label": "正在停止，等待已发出的识别结束…" if stopping
                        else f"已完成 {completed}/{total}，{workers} 路并发识别中…",
                    "done": completed, "total": total, "current_id": doc_id,
                    "ok": bool(items_by_id[doc_id]["ok"]), "error": error, "workers": workers,
                })

        was_cancelled = cancel_event.is_set()
        emit({
            "event": "aidoc_recognize_batch", "job_id": job_id,
            "stage": "cancelled" if was_cancelled else "done",
            "stage_label": f"已取消，完成 {completed - cancelled} 份" if was_cancelled else "批量识别完成",
            "done": completed, "total": total, "workers": workers,
        })
        return {
            "job_id": job_id, "total": total, "done": completed, "workers": workers,
            "ai_ok": ai_ok, "ai_fail": ai_fail, "cancelled": cancelled,
            "was_cancelled": was_cancelled,
            "items": [items_by_id[doc_id] for doc_id in ids],
        }
    finally:
        with RECOGNITION_JOBS_LOCK:
            RECOGNITION_JOBS.pop(job_id, None)


def command_cancel_recognition(args: dict[str, Any]) -> dict[str, Any]:
    """停止尚未开始的识别；已经发给 AI 的最多 workers 份会完成后退出。"""
    require_unlocked()
    job_id = str(args.get("job_id") or "").strip()
    with RECOGNITION_JOBS_LOCK:
        job = RECOGNITION_JOBS.get(job_id)
        if not job:
            return {"job_id": job_id, "found": False, "cancelled_pending": 0}
        job["cancel"].set()
        futures: list[Future[Any]] = list(job.get("futures") or [])
        cancelled_pending = sum(1 for future in futures if future.cancel())
    return {"job_id": job_id, "found": True, "cancelled_pending": cancelled_pending}


def _row_to_suggestion(r: sqlite3.Row) -> dict[str, Any]:
    """DB 行 → 统一 suggestion 结构（给 project_to_profile 用）。"""
    try:
        extra = json.loads(r["extra_fields"]) if r["extra_fields"] else {}
    except Exception:
        extra = {}
    dtype = r["document_type"] or "other"
    return {
        "document_type": dtype,
        "document_type_label": r["document_type_label"] or docintel.DOCUMENT_TYPES.get(dtype, "其他资料"),
        "company_name": r["company_name"] or "",
        "brand": r["brand"] or "",
        "certificate_no": r["certificate_no"] or "",
        "issuer": r["issuer"] or "",
        "issued_at": r["issued_at"] or "",
        "expires_at": r["expires_at"] or "",
        "applicable_scope": r["applicable_scope"] or "",
        "extra_fields": extra if isinstance(extra, dict) else {},
        "ai_summary": r["ai_summary"] or "",
        "ai_confidence": int(r["ai_confidence"] or 0),
    }


def command_document_meta(_: dict[str, Any]) -> dict[str, Any]:
    """证件类型、主列表列、各类型版面字段——前端动态渲染用（对齐 版面样式.xlsx）。"""
    return {
        "types": [{"key": k, "label": v} for k, v in docintel.DOCUMENT_TYPES.items()],
        "list_columns": docintel.LIST_COLUMNS,
        "profiles": docintel.FIELD_PROFILES,
        "default_profile": docintel.DEFAULT_PROFILE,
    }


def command_unrecognized_documents(args: dict[str, Any]) -> dict[str, Any]:
    """正式识别前返回全部未识别队列，不受前端当前分页限制。"""
    require_unlocked()
    limit = max(1, min(5000, int(args.get("limit") or 5000)))
    conn = connect_db(library_dir())
    rows = conn.execute(
        "SELECT id, file_name, page_count FROM documents WHERE recognized_at IS NULL ORDER BY id ASC LIMIT ?",
        (limit,),
    ).fetchall()
    conn.close()
    return {
        "items": [
            {"id": int(row["id"]), "file_name": str(row["file_name"]), "page_count": int(row["page_count"] or 1)}
            for row in rows
        ],
    }


def command_list_documents(args: dict[str, Any]) -> dict[str, Any]:
    """资料库列表：按主列表列返回，支持按 review_status / document_type 过滤。"""
    require_unlocked()
    root = library_dir()
    conn = connect_db(root)
    where, params = ["1=1"], []
    status = str(args.get("review_status") or "").strip()
    dtype = str(args.get("document_type") or "").strip()
    only_unrecognized = bool(args.get("only_unrecognized"))
    if status:
        where.append("review_status = ?")
        params.append(status)
    if only_unrecognized:
        where.append("recognized_at IS NULL")
    if dtype:
        where.append("document_type = ?")
        params.append(dtype)
    page = max(1, int(args.get("page") or 1))
    per_page = max(5, min(100, int(args.get("per_page") or 20)))
    filtered_total = int(conn.execute(
        f"SELECT COUNT(*) FROM documents WHERE {' AND '.join(where)}", params
    ).fetchone()[0])
    total_pages = max(1, (filtered_total + per_page - 1) // per_page)
    page = min(page, total_pages)
    offset = (page - 1) * per_page
    rows = conn.execute(
        f"SELECT * FROM documents WHERE {' AND '.join(where)} ORDER BY id DESC LIMIT ? OFFSET ?",
        [*params, per_page, offset],
    ).fetchall()

    docs = []
    for r in rows:
        sug = _row_to_suggestion(r)
        list_values = {}
        for col in docintel.LIST_COLUMNS:
            key = col["key"]
            list_values[key] = sug["document_type_label"] if key == "document_type_label" else sug.get(key, "")
        # 完整字段值（顶层 + extra:*），前端按所选类型动态切换列时用 values[source] 取值。
        values = {
            "document_type_label": sug["document_type_label"],
            "company_name": sug.get("company_name", ""), "brand": sug.get("brand", ""),
            "certificate_no": sug.get("certificate_no", ""), "issuer": sug.get("issuer", ""),
            "issued_at": sug.get("issued_at", ""), "expires_at": sug.get("expires_at", ""),
            "applicable_scope": sug.get("applicable_scope", ""),
        }
        for k, v in (sug.get("extra_fields") or {}).items():
            values["extra:" + str(k)] = v
        docs.append({
            "id": r["id"], "file_name": r["file_name"], "extension": r["extension"],
            "page_count": r["page_count"], "review_status": r["review_status"] or "",
            "recognized": bool(r["recognized_at"]), "is_duplicate": bool(r["duplicate_of_id"]),
            "document_type": sug["document_type"], "document_type_label": sug["document_type_label"],
            "ai_confidence": sug["ai_confidence"], "list_values": list_values, "values": values,
            "managed_path": r["managed_path"] or "", "original_path": r["original_path"] or "",
        })

    def count(sql, p=()):
        return int(conn.execute(sql, p).fetchone()[0])

    stats = {
        "total": count("SELECT COUNT(*) FROM documents"),
        "unrecognized": count("SELECT COUNT(*) FROM documents WHERE recognized_at IS NULL"),
        "pending": count("SELECT COUNT(*) FROM documents WHERE review_status='pending_review'"),
        "confirmed": count("SELECT COUNT(*) FROM documents WHERE review_status='confirmed'"),
        "duplicate": count("SELECT COUNT(*) FROM documents WHERE duplicate_of_id IS NOT NULL"),
    }
    conn.close()
    return {
        "documents": docs,
        "stats": stats,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": filtered_total,
            "total_pages": total_pages,
        },
    }


def command_get_document(args: dict[str, Any]) -> dict[str, Any]:
    """单份资料详情：按类型版面投影出有序字段，给详情/编辑用。"""
    require_unlocked()
    doc_id = int(args.get("id") or 0)
    conn = connect_db(library_dir())
    r = conn.execute("SELECT * FROM documents WHERE id = ?", (doc_id,)).fetchone()
    conn.close()
    if not r:
        raise RuntimeError("资料不存在。")
    sug = _row_to_suggestion(r)
    return {
        "id": r["id"], "file_name": r["file_name"], "extension": r["extension"],
        "page_count": r["page_count"], "managed_path": r["managed_path"] or "", "original_path": r["original_path"] or "",
        "review_status": r["review_status"] or "", "recognized": bool(r["recognized_at"]),
        "document_type": sug["document_type"], "document_type_label": sug["document_type_label"],
        "ai_confidence": sug["ai_confidence"], "ai_summary": sug["ai_summary"],
        "profile": docintel.project_to_profile(sug),
    }


# 干净的类型短名（用于文件夹和规范文件名，对齐「版面样式.xlsx」表名）
TYPE_FOLDER_NAMES = {
    "business_license": "营业执照", "trademark_certificate": "商标证",
    "authorization_letter": "授权书", "barcode_certificate": "条码证",
    "quality_report": "质检报告", "ccc_certificate": "3C证书",
    "production_license": "生产许可证", "hygiene_license": "卫生许可证",
    "product_filing": "产品备案", "food_license": "食品许可证",
    "contract": "合同", "tax_certificate": "税务资料", "design_drawing": "设计稿图纸", "other": "其他资料",
}


def _type_folder(dtype: str) -> str:
    name = TYPE_FOLDER_NAMES.get(dtype)
    if name:
        return name
    return re.sub(r'[\\/:*?"<>|]+', "", docintel.DOCUMENT_TYPES.get(dtype, "其他资料")).strip() or "其他资料"


def _safe_name(text: Any, limit: int = 40) -> str:
    """去掉文件系统非法字符 + 空白，截断。"""
    s = re.sub(r'[\\/:*?"<>|\r\n\t]+', "", str(text or "")).strip().strip(".")
    return s[:limit]


def _standard_filename(row: sqlite3.Row, ext: str) -> str:
    """规范文件名：公司_类型_编号（编号缺则用到期日/起始日）。缺失项自动跳过。
    例：西安稻叶山供应链管理有限公司_合同_DYS20240418.pdf"""
    company = _safe_name(row["company_name"], 30)
    label = _safe_name(_type_folder(row["document_type"] or "other"), 12)
    ident = _safe_name(row["certificate_no"] or row["expires_at"] or row["issued_at"], 30)
    parts = [p for p in (company, label, ident) if p]
    base = "_".join(parts) or _safe_name(Path(str(row["file_name"] or "资料")).stem, 60) or "资料"
    return base[:90] + (ext or "")


def _organize_file(conn: sqlite3.Connection, doc_id: int, dtype: str, issued_at: str) -> str:
    """确认入库后：把文件【规范重命名】并放进「库根/年份/类型/公司_类型_编号.ext」，更新 managed_path。
    - 有库内副本(复制管理) → 移动并重命名；
    - 没有副本(仅建索引) → 复制原文件进库归类（不动用户原文件）。
    年份取证件日期(issued_at)，认不出用当前年。返回新路径，没法组织返回空。"""
    row = conn.execute("SELECT * FROM documents WHERE id = ?", (doc_id,)).fetchone()
    if not row:
        return ""
    managed = str(row["managed_path"] or "").strip()
    original = str(row["original_path"] or "").strip()
    src = None
    move_it = False
    if managed and Path(managed).exists():
        src = Path(managed)
        move_it = True             # 复制管理：移动已有副本
    elif original and Path(original).exists():
        src = Path(original)
        move_it = False            # 仅建索引/副本丢失：复制原文件进库，不动原件
    if src is None:
        return ""
    root = library_dir()
    m = re.match(r"\s*(\d{4})", str(issued_at or ""))
    year = m.group(1) if (m and 1990 <= int(m.group(1)) <= 2100) else str(datetime.now().year)
    dest_dir = root / year / _type_folder(dtype)
    dest_dir.mkdir(parents=True, exist_ok=True)
    target_name = _standard_filename(row, src.suffix.lower())
    stem, suf = Path(target_name).stem, Path(target_name).suffix
    dest = dest_dir / target_name
    i = 2
    while dest.exists() and dest != src:
        dest = dest_dir / f"{stem}({i}){suf}"
        i += 1
    if dest == src:
        return managed or str(dest)
    try:
        if move_it:
            shutil.move(str(src), str(dest))
        else:
            shutil.copy2(str(src), str(dest))
    except Exception:
        return ""
    conn.execute("UPDATE documents SET managed_path = ?, file_name = ? WHERE id = ?", (str(dest), dest.name, doc_id))
    return str(dest)


def command_confirm_document(args: dict[str, Any]) -> dict[str, Any]:
    """保存人工核对后的字段并改状态（确认/驳回）。确认时把文件归类移动到 年份/类型/ 文件夹。
    args: {id, document_type, values:{source: value}, review_status}"""
    require_unlocked()
    doc_id = int(args.get("id") or 0)
    if doc_id <= 0:
        raise RuntimeError("缺少资料 id。")
    dtype = str(args.get("document_type") or "other")
    if dtype not in docintel.DOCUMENT_TYPES:
        dtype = "other"
    review_status = str(args.get("review_status") or "confirmed")
    values = args.get("values") if isinstance(args.get("values"), dict) else {}

    top = {"company_name": "", "brand": "", "certificate_no": "", "issuer": "",
           "issued_at": "", "expires_at": "", "applicable_scope": ""}
    extra: dict[str, Any] = {}
    for source, value in values.items():
        text = str(value or "").strip()
        if str(source).startswith("extra:"):
            key = str(source).split(":", 1)[1]
            if text:
                extra[key] = text
        elif source in top:
            top[source] = docintel.normalize_date(text) if source in ("issued_at", "expires_at") else text

    conn = connect_db(library_dir())
    now = datetime.now().isoformat(timespec="seconds")
    conn.execute(
        """
        UPDATE documents SET
          document_type = ?, document_type_label = ?, company_name = ?, brand = ?,
          certificate_no = ?, issuer = ?, issued_at = ?, expires_at = ?, applicable_scope = ?,
          extra_fields = ?, review_status = ?, updated_at = ?
        WHERE id = ?
        """,
        (dtype, docintel.DOCUMENT_TYPES[dtype], top["company_name"], top["brand"], top["certificate_no"],
         top["issuer"], top["issued_at"], top["expires_at"], top["applicable_scope"],
         json.dumps(extra, ensure_ascii=False), review_status, now, doc_id),
    )
    conn.commit()
    duplicate_of = _mark_duplicate_by_cert(conn, doc_id, top["certificate_no"])
    # 确认入库 → 把文件移动到 年份/类型/ 文件夹归类
    moved_to = ""
    if review_status == "confirmed":
        moved_to = _organize_file(conn, doc_id, dtype, top["issued_at"])
        try:
            templatepool.upsert_from_document(conn, doc_id, active=True)
        except Exception:
            pass
    conn.commit()
    conn.close()
    schedule_knowledge_index(doc_id, force=True)
    return {"id": doc_id, "review_status": review_status, "duplicate_of_id": duplicate_of, "moved_to": moved_to}


def command_delete_document(args: dict[str, Any]) -> dict[str, Any]:
    """永久删除资料文件及数据库记录。

    删除边界：
    - 存在 managed_path 时，只删除资料库管理的文件，不碰导入前的库外原件；
    - 仅建索引且没有 managed_path 时，删除 original_path 指向的原文件；
    - 文件删除失败时保留数据库记录，避免界面显示已删但磁盘文件仍存在。
    """
    require_unlocked()
    doc_id = int(args.get("id") or 0)
    if doc_id <= 0:
        raise RuntimeError("缺少资料 id。")

    root = library_dir().resolve()
    conn = connect_db(root)
    row = conn.execute(
        "SELECT id, file_name, original_path, managed_path, import_mode FROM documents WHERE id = ?",
        (doc_id,),
    ).fetchone()
    if not row:
        conn.close()
        raise RuntimeError("资料不存在或已经删除。")

    managed_text = str(row["managed_path"] or "").strip()
    original_text = str(row["original_path"] or "").strip()
    import_mode = str(row["import_mode"] or "copy")
    target: Path | None = None
    deleted_scope = "record_only"

    if managed_text:
        managed = Path(managed_text).expanduser().resolve()
        # managed_path 按设计只能指向资料库。数据库异常时拒绝越界删除。
        if not managed.is_relative_to(root):
            conn.close()
            raise RuntimeError("资料文件路径不在本地资料库内，已拒绝删除。")
        target = managed
        deleted_scope = "managed_copy"
    elif import_mode == "index" and original_text:
        target = Path(original_text).expanduser().resolve()
        deleted_scope = "indexed_original"

    file_deleted = False
    if target is not None and target.exists():
        if not target.is_file():
            conn.close()
            raise RuntimeError("目标路径不是文件，已拒绝删除。")
        try:
            target.unlink()
            file_deleted = True
        except OSError as exc:
            conn.close()
            raise RuntimeError(f"文件删除失败，可能正被其他程序占用：{exc}") from exc

    # 清理其它资料对本记录的“重复文件”引用，再删除主记录。
    conn.execute(
        "UPDATE documents SET duplicate_of_id = NULL, duplicate_reason = NULL WHERE duplicate_of_id = ?",
        (doc_id,),
    )
    conn.execute("DELETE FROM documents WHERE id = ?", (doc_id,))
    conn.commit()
    conn.close()
    return {
        "id": doc_id,
        "file_name": str(row["file_name"] or ""),
        "file_deleted": file_deleted,
        "deleted_scope": deleted_scope,
    }


# ───────────────────────── AI 资料员：按需整理材料 ─────────────────────────
# “脑细胞”种子：办事关键词 → 通常需要的资料类型。以后服务器可下发更新，越来越聪明。
TASK_MATERIALS = {
    "投标": ["business_license", "trademark_certificate", "authorization_letter", "quality_report", "ccc_certificate"],
    "招标": ["business_license", "authorization_letter", "quality_report"],
    "入驻": ["business_license", "trademark_certificate", "authorization_letter", "quality_report"],
    "开店": ["business_license", "trademark_certificate", "food_license"],
    "认证": ["ccc_certificate", "quality_report", "product_filing"],
    "授权": ["authorization_letter", "trademark_certificate", "business_license"],
    "报关": ["business_license", "quality_report"],
    "供应商": ["business_license", "trademark_certificate", "authorization_letter", "quality_report"],
    "资质": ["business_license", "quality_report", "ccc_certificate", "production_license"],
    "营业执照": ["business_license"],
    "合同": ["contract"],
    "商标": ["trademark_certificate"],
    "质检": ["quality_report"],
}


def command_find_materials(args: dict[str, Any]) -> dict[str, Any]:
    """AI 资料员 v1：用户描述要办的事 → 本地关键词 + 办事→材料种子规则 匹配库里资料。
    返回按相关度排序的候选（以后可换服务器 AI 脑细胞）。args: {query}"""
    require_unlocked()
    query = str(args.get("query") or "").strip()
    conn = connect_db(library_dir())
    rows = conn.execute("SELECT * FROM documents ORDER BY id DESC").fetchall()
    conn.close()
    words = [w for w in re.split(r"[\s,，。、;；/()（）]+", query) if len(w) >= 1]
    needed_types: set[str] = set()
    for kw, types in TASK_MATERIALS.items():
        if kw in query:
            needed_types.update(types)
    results = []
    for r in rows:
        sug = _row_to_suggestion(r)
        hay = " ".join(str(x or "") for x in [
            sug["document_type_label"], sug["company_name"], sug["brand"],
            sug["certificate_no"], r["file_name"], sug["applicable_scope"]])
        score = 0
        reasons = []
        if sug["document_type"] in needed_types:
            score += 3
            reasons.append("办这事通常需要「%s」" % sug["document_type_label"])
        for w in words:
            if w and w in hay:
                score += 1
                reasons.append("含「%s」" % w)
        if not query:
            score = 1  # 没填描述就列全部，让用户手动挑
        if score > 0:
            results.append({
                "id": r["id"], "file_name": r["file_name"],
                "type_label": sug["document_type_label"], "company": sug["company_name"],
                "score": score, "reason": "；".join(dict.fromkeys(reasons))[:80],
            })
    results.sort(key=lambda x: -x["score"])
    return {
        "query": query,
        "results": results,
        "needed_types": [docintel.DOCUMENT_TYPES.get(t, t) for t in needed_types],
    }


def _contract_context_payload(value: Any) -> dict[str, Any]:
    """只保留续写合同所需字段，避免把整个聊天消息重新发送。"""
    if not isinstance(value, dict):
        return {}
    items = []
    for raw in list(value.get("line_items") or [])[:30]:
        if not isinstance(raw, dict) or not str(raw.get("name") or "").strip():
            continue
        items.append({
            "name": str(raw.get("name") or "").strip()[:200],
            "specification": str(raw.get("specification") or "").strip()[:200],
            "quantity": raw.get("quantity") or 0,
            "unit": str(raw.get("unit") or "").strip()[:20],
            "unit_price": raw.get("unit_price") or 0,
            "remark": str(raw.get("remark") or "").strip()[:200],
        })
    context = {
        "template_id": int(value.get("template_id") or 0),
        "source_document_id": int(value.get("source_document_id") or 0),
        "supplier_name": str(value.get("supplier_name") or "").strip()[:200],
        "payee_name": str(value.get("payee_name") or "").strip()[:200],
        "bank_name": str(value.get("bank_name") or "").strip()[:200],
        "bank_account": re.sub(r"\s+", "", str(value.get("bank_account") or "").strip())[:100],
        "supporting_ids": [int(item) for item in list(value.get("supporting_ids") or [])[:20] if str(item).isdigit()],
        "line_items": items,
        "delivery_address": str(value.get("delivery_address") or "").strip()[:300],
        "delivery_deadline": str(value.get("delivery_deadline") or "").strip()[:300],
        "payment_terms": str(value.get("payment_terms") or "").strip()[:500],
        "output_name": str(value.get("output_name") or "新合同").strip()[:100],
    }
    if context["template_id"] <= 0 or not context["supplier_name"] or not items:
        return {}
    return context


def _assistant_needs_template_context(message: str) -> bool:
    text = str(message or "")
    action = re.search(r"生成|起草|拟定|拟一|写一|制作|导出|新建|修改|替换|套用", text)
    document = re.search(r"合同|协议|授权书|证明|声明|申请书|文书", text)
    return bool(action and document)


def _assistant_local_clarification(message: str) -> dict[str, Any] | None:
    """极短且含糊的需求先澄清，不把整个资料库压给远端模型硬猜。"""
    text = re.sub(r"\s+", "", str(message or "")).strip("。！？!?，,")
    if text in {"你好", "您好", "在吗", "小秘书", "嗨"}:
        return {
            "reply": "我在。你这次想查资料、问资料内容，还是生成文书？",
            "quick_options": [
                {"label": "查找资料", "message": "我想查找一份资料"},
                {"label": "询问内容", "message": "我想询问资料里的具体内容"},
                {"label": "生成文书", "message": "我想生成一份文书"},
            ],
        }
    if re.fullmatch(r"(?:帮我|麻烦)?(?:做|写|弄|处理|看)?(?:一下|一份|一个)?合同", text):
        return {
            "reply": "可以。你是要查找已有合同、起草新合同，还是修改库存中的合同？",
            "quick_options": [
                {"label": "查找已有合同", "message": "请帮我查找现有合同"},
                {"label": "起草新合同", "message": "我要起草一份新合同，请告诉我需要提供哪些信息"},
                {"label": "修改库存合同", "message": "我要基于库存中的合同修改并生成新合同"},
                {"label": "检查合同内容", "message": "我要询问或检查一份合同的具体内容"},
            ],
        }
    if re.fullmatch(r"(?:帮我|麻烦)?(?:找|查|整理|处理|看)?(?:一下|一些)?(?:资料|文件|材料|文档)", text):
        return {
            "reply": "可以。你希望按什么方式处理这些资料？",
            "quick_options": [
                {"label": "按名称查找", "message": "我要按文件名或关键词查找资料"},
                {"label": "按公司查找", "message": "我要查找某家公司的全部资料"},
                {"label": "检查缺失材料", "message": "我要检查办理事项还缺哪些材料"},
                {"label": "整理导出", "message": "我要把相关资料整理到一个文件夹"},
            ],
        }
    return None


def _assistant_fallback(message: str, error: str) -> dict[str, Any]:
    return {
        "reply": "本轮连接在回答完成前中断了。先不让对话卡死，你可以直接重试，或者先缩小处理范围。",
        "materials": [],
        "need_follow_up": True,
        "missing_materials": [],
        "watermark_text": "",
        "contract_job": None,
        "document_job": None,
        "model": "",
        "training_notes_used": [],
        "templates_available": 0,
        "quick_options": [
            {"label": "重试刚才问题", "message": str(message or "")[:2000]},
            {"label": "先查相关资料", "message": "先帮我查找与刚才问题相关的资料，不要生成文件"},
            {"label": "我来补充条件", "message": "请先列出需要我补充的关键信息，不要直接生成文件"},
        ],
        "rateable": False,
        "response_kind": "connection_fallback",
        "connection_error": str(error or "")[:300],
    }


def command_assistant_task_create(args: dict[str, Any]) -> dict[str, Any]:
    """Accept spreadsheet files into the secretary's local task inbox."""
    require_unlocked()
    root = library_dir()
    task = assistanttasks.create_task(
        db_path(root),
        root,
        [str(value) for value in list(args.get("paths") or [])],
        str(args.get("instruction") or "").strip(),
        str(args.get("conversation_id") or "").strip(),
    )
    return {"task": task}


def command_assistant_tasks_list(args: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    root = library_dir()
    assistanttasks.resume_pending(db_path(root))
    conn = connect_db(root)
    try:
        tasks = assistanttasks.list_tasks(conn, int(args.get("limit") or 100))
    finally:
        conn.close()
    active = sum(item["status"] in {"received", "processing"} for item in tasks)
    ready = sum(item["status"] == "ready" for item in tasks)
    return {"tasks": tasks, "active": active, "ready": ready, "dir": str(root / ".assistant" / "tasks")}


def command_assistant_task_get(args: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    conn = connect_db(library_dir())
    try:
        task = assistanttasks.get_task(conn, str(args.get("id") or ""))
    finally:
        conn.close()
    return {"task": task}


def command_assistant_task_archive(args: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    conn = connect_db(library_dir())
    try:
        task = assistanttasks.archive_task(conn, str(args.get("id") or ""))
    finally:
        conn.close()
    return {"task": task}


def command_assistant_task_organize(args: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    return assistanttasks.organize_product_catalog(db_path(library_dir()), str(args.get("id") or ""))


def _assistant_local_task_reply(message: str, history: list[Any] | None = None) -> dict[str, Any] | None:
    """Answer common task-status questions locally so they never wait on the network AI."""
    task_reference = bool(re.search(r"(?:托管|表格|Excel|excel|CSV|csv|那个表|这张表|这表|的表|货盘表)", message))
    export_shortcut = bool(re.fullmatch(r"\s*(?:导出|生成)(?:为)?\s*Excel\s*", message, re.I))
    if not task_reference and not export_shortcut:
        return None
    conn = connect_db(library_dir())
    try:
        tasks = assistanttasks.list_tasks(conn, 20)
    finally:
        conn.close()
    if not tasks:
        return None
    words = [word for word in re.split(r"[\s，。！？、,;；]+", message) if len(word) >= 2]
    task = max(
        tasks,
        key=lambda item: (
            sum(word.lower() in (item["title"] + " " + item["instruction"] + " " + " ".join(
                file["original_name"] for file in item["files"]
            )).lower() for word in words),
            item["updated_at"],
        ),
    )
    conn = connect_db(library_dir())
    try:
        task = assistanttasks.get_task(conn, task["id"])
    finally:
        conn.close()
    context_parts = [message]
    for item in list(history or [])[-8:]:
        if isinstance(item, dict) and str(item.get("role") or "") in {"user", "human"}:
            context_parts.append(str(item.get("content") or item.get("text") or ""))
    request_context = "\n".join(context_parts)
    requested_fields = [
        field for field in ("商品名称", "品牌", "型号", "编码", "条码", "电商链接", "销售价格", "代发价格")
        if field in request_context
    ]
    organize_requested = bool(re.search(
        r"(?:整理整齐|整理出来|按.+(?:字段|栏目|列)|只保留|提取.+字段|导出(?:为)?\s*Excel|生成.+Excel)",
        message,
        re.I | re.S,
    )) or export_shortcut
    if organize_requested and (len(requested_fields) >= 2 or export_shortcut):
        result = assistanttasks.organize_product_catalog(db_path(library_dir()), task["id"])
        return {
            "reply": (
                f"已经真正生成整理版 Excel，共提取 {result['rows']} 条商品记录。"
                "只保留商品名称、品牌、型号、编码、条码、电商链接、销售价格、代发价格；原表没有修改。"
            ),
            "materials": [], "need_follow_up": False, "missing_materials": [], "watermark_text": "",
            "contract_job": None, "document_job": None, "generated_spreadsheet": {
                "file_path": result["file_path"], "dir": result["dir"], "file_name": result["file_name"],
                "rows": result["rows"], "missing": result["missing"],
            },
            "model": "local-sheet-organizer", "training_notes_used": [], "templates_available": 0,
            "quick_options": [], "rateable": True, "response_kind": "spreadsheet_output",
        }
    if organize_requested and len(requested_fields) < 2:
        return {
            "reply": "可以真正整理并生成一个新 Excel，但我需要先知道保留哪些字段。原表会保持不动。",
            "materials": [], "need_follow_up": True, "missing_materials": [], "watermark_text": "",
            "contract_job": None, "document_job": None, "generated_spreadsheet": None,
            "model": "local-sheet-organizer", "training_notes_used": [], "templates_available": 0,
            "quick_options": [{
                "label": "整理商品货盘",
                "message": "请把这份表按商品名称、品牌、型号、编码、条码、电商链接、销售价格、代发价格整理出来，其他列不要",
            }],
            "rateable": True, "response_kind": "spreadsheet_clarification",
        }
    if not re.search(r"(?:怎么样|咋样|好了吗|完成了吗|进度|结果|梳理|处理到哪|看完了吗|刚才|之前|整理)", message):
        return None
    status = task["status"]
    generated_spreadsheet = None
    if status == "ready":
        reply = f"这件表格事项已经梳理好了：\n\n{task['summary']}"
        options = [
            {"label": "继续查异常", "message": f"继续分析托管事项“{task['title']}”里需要我注意的异常"},
            {"label": "暂时不用", "message": "好的，先放在托管事项里，我之后再处理"},
        ]
    elif status == "complete":
        output = next((item for item in task["files"] if item.get("role") == "output"), None)
        reply = task["summary"] or f"“{task['title']}”已经整理完成。"
        options = []
        if output:
            generated_spreadsheet = {
                "file_path": output["stored_path"], "dir": str(Path(output["stored_path"]).parent),
                "file_name": output["original_name"], "rows": int(
                    ((task.get("analysis") or {}).get("organization") or {}).get("rows") or 0
                ),
                "missing": ((task.get("analysis") or {}).get("organization") or {}).get("missing") or {},
            }
    elif status == "error":
        reply = f"这件表格事项处理失败了：{task['error']}\n原表没有被修改，你可以重新托管或换成 xlsx/csv 再试。"
        options = []
    elif status == "archived":
        reply = f"“{task['title']}”已经归档。\n\n{task['summary'] or '当时没有生成梳理摘要。'}"
        options = []
    else:
        reply = f"“{task['title']}”还在本机后台梳理，原表没有被修改。你可以先忙别的，完成后会显示为“待你查看”。"
        options = [{"label": "查看托管事项", "message": "打开表格托管事项看看当前进度"}]
    return {
        "reply": reply, "materials": [], "need_follow_up": status not in {"ready", "error", "archived"},
        "missing_materials": [], "watermark_text": "", "contract_job": None, "document_job": None,
        "model": "local-task-inbox", "training_notes_used": [], "templates_available": 0,
        "quick_options": options, "rateable": True, "response_kind": "managed_task",
        "generated_spreadsheet": generated_spreadsheet,
    }


def command_assistant_chat(args: dict[str, Any]) -> dict[str, Any]:
    """把精简库存和对话发给服务器 AI 资料员；原文件始终留在本机。"""
    require_unlocked()
    message = str(args.get("message") or "").strip()
    if not message:
        raise RuntimeError("请输入要办理的事项。")
    local_task_reply = _assistant_local_task_reply(message, list(args.get("history") or []))
    if local_task_reply:
        return local_task_reply
    token = str(STATE.get("token") or "")
    if not token:
        raise RuntimeError("软件尚未在线登记，无法使用 AI 档案秘书。请先检查网络并重新登记。")

    clarification = _assistant_local_clarification(message)
    if clarification:
        return {
            **clarification,
            "materials": [], "need_follow_up": True, "missing_materials": [],
            "watermark_text": "", "contract_job": None, "document_job": None,
            "model": "local-dialog-router", "training_notes_used": [],
            "templates_available": 0, "rateable": True, "response_kind": "clarification",
        }

    conn = connect_db(library_dir())
    rows = list(conn.execute(
        """
        SELECT id, file_name, document_type, document_type_label, company_name, brand,
               issued_at, expires_at, applicable_scope, review_status, recognized_at
        FROM documents
        ORDER BY id DESC
        LIMIT 2000
        """
    ).fetchall())
    training_notes = _matching_training_notes(conn, message)
    use_template_context = _assistant_needs_template_context(message)
    template_rows = conn.execute(
        "SELECT * FROM document_templates WHERE enabled = 1 ORDER BY updated_at DESC LIMIT 300"
    ).fetchall() if use_template_context else []
    template_rows = sorted(
        template_rows,
        key=lambda row: (templatepool.rank_for_message(row, message), int(row["id"])),
        reverse=True,
    )[:6]
    template_pool = []
    content_budget = 36000
    for row in template_rows:
        item = templatepool.payload(row)
        body = str(item.get("template_text") or item.get("content_text") or "")
        body = body[: min(18000, content_budget)]
        content_budget -= len(body)
        item["template_text"] = body
        item.pop("content_text", None)
        template_pool.append(item)
        if content_budget <= 0:
            break
    try:
        knowledge_hits = knowledge.search(conn, message, limit=12)
    except Exception as exc:
        _knowledge_log(f"search error={type(exc).__name__}:{str(exc)[:240]}")
        knowledge_hits = []
    managed_tasks = assistanttasks.recent_context(conn)
    conn.close()

    query_words = [w for w in re.split(r"[\s,，。、;；/()（）]+", message) if w]
    needed_types: set[str] = set()
    knowledge_ids = {int(item.get("document_id") or 0) for item in knowledge_hits}
    for keyword, material_types in TASK_MATERIALS.items():
        if keyword in message:
            needed_types.update(material_types)

    def inventory_rank(row: sqlite3.Row) -> tuple[int, int]:
        haystack = " ".join(str(row[key] or "") for key in (
            "file_name", "document_type_label", "company_name", "brand", "applicable_scope"
        ))
        score = 3 if str(row["document_type"] or "") in needed_types else 0
        if int(row["id"]) in knowledge_ids:
            score += 20
        score += sum(2 for word in query_words if word and word in haystack)
        if str(row["review_status"] or "") == "confirmed":
            score += 1
        return score, int(row["id"])

    # 控制上下文体积：相关资料优先，其次是最近确认过的资料，最多发 300 条元数据。
    rows.sort(key=inventory_rank, reverse=True)
    rows = rows[:180]

    inventory: list[dict[str, Any]] = []
    local: dict[int, dict[str, Any]] = {}
    for row in rows:
        item = {
            "id": int(row["id"]),
            "file_name": str(row["file_name"] or ""),
            "document_type": str(row["document_type"] or ""),
            "type_label": str(row["document_type_label"] or ""),
            "company": str(row["company_name"] or ""),
            "brand": str(row["brand"] or ""),
            "issued_at": str(row["issued_at"] or ""),
            "expires_at": str(row["expires_at"] or ""),
            "scope": str(row["applicable_scope"] or "")[:200],
            "review_status": str(row["review_status"] or ""),
        }
        inventory.append(item)
        local[item["id"]] = item

    history: list[dict[str, str]] = []
    contract_context: dict[str, Any] = {}
    for entry in list(args.get("history") or [])[-12:]:
        if not isinstance(entry, dict):
            continue
        role = "assistant" if str(entry.get("role") or "") in {"assistant", "ai"} else "user"
        content = str(entry.get("content") or "").strip()
        if content:
            history.append({"role": role, "content": content[:3000]})
        if role == "assistant":
            candidate = _contract_context_payload(
                entry.get("contract_context") or entry.get("contractContext")
            )
            if candidate:
                contract_context = candidate

    try:
        response = request_json(
            "/api/desktop/assistant/chat",
            {
                "message": message[:2000],
                "history": history,
                "contract_context": contract_context or None,
                "knowledge_hits": knowledge_hits,
                "inventory": inventory,
                "need_organize": bool(args.get("need_organize")),
                "use_watermark": bool(args.get("use_watermark")),
                "watermark_text": str(args.get("watermark_text") or "")[:100],
                "training_notes": training_notes,
                "template_pool": template_pool,
                "managed_tasks": managed_tasks,
            },
            token=token,
            timeout=180,
        )
    except urllib.error.HTTPError as exc:
        try:
            detail = json.loads(exc.read().decode("utf-8", "replace"))
            error = str(detail.get("message") or detail)
        except Exception:
            error = str(exc)
        if int(exc.code) >= 500:
            return _assistant_fallback(message, f"HTTP {exc.code}: {error[:240]}")
        raise RuntimeError(f"AI 档案秘书请求失败（HTTP {exc.code}）：{error[:300]}") from exc
    except Exception as exc:
        return _assistant_fallback(message, f"{type(exc).__name__}: {str(exc)[:240]}")

    ids = []
    for value in response.get("gather_ids") or []:
        try:
            doc_id = int(value)
        except (TypeError, ValueError):
            continue
        if doc_id in local and doc_id not in ids:
            ids.append(doc_id)
    reasons = response.get("selection_reasons") if isinstance(response.get("selection_reasons"), dict) else {}
    materials = []
    for doc_id in ids:
        item = local[doc_id]
        materials.append({
            "id": doc_id,
            "file_name": item["file_name"],
            "type_label": item["type_label"] or item["document_type"] or "未分类",
            "company": item["company"],
            "reason": str(reasons.get(str(doc_id)) or reasons.get(doc_id) or "AI 档案秘书推荐")[:200],
        })

    contract_job = response.get("contract_job") if isinstance(response.get("contract_job"), dict) else None
    if contract_job:
        template_id = int(contract_job.get("template_id") or 0)
        matching_template = next(
            (item for item in template_pool if int(item.get("id") or 0) == template_id), None
        )
        if matching_template:
            contract_job["source_document_id"] = int(matching_template.get("source_document_id") or 0)
    document_job = response.get("document_job") if isinstance(response.get("document_job"), dict) else None
    reply = str(response.get("reply") or "我已经核对了当前库存。")
    if contract_job or document_job:
        reply = re.sub(r"已(?:经)?(?:生成|导出|制作)", "已准备生成", reply)
    if not contract_job and not document_job and re.search(
        r"已(?:经)?(?:生成|导出|制作).{0,40}(?:合同|文件|文书|PDF|DOCX)", reply, re.I,
    ):
        reply += "\n\n系统核验：AI 没有下发文件生成任务，本机尚未生成文件。请补充关键信息并明确要求生成 PDF。"

    return {
        "reply": reply,
        "materials": materials,
        "need_follow_up": bool(response.get("need_follow_up")),
        "missing_materials": list(response.get("missing_materials") or []),
        "watermark_text": str(response.get("watermark_text") or "")[:60],
        "contract_job": contract_job,
        "document_job": document_job,
        "model": str(response.get("model") or ""),
        "training_notes_used": [note["title"] for note in training_notes],
        "templates_available": len(template_pool),
        "quick_options": list(response.get("quick_options") or [])[:4]
            if isinstance(response.get("quick_options"), list) else [],
        "rateable": True,
        "response_kind": "answer",
    }


def command_generate_contract(args: dict[str, Any]) -> dict[str, Any]:
    """执行服务器给出的合同生成任务，输出一个全新的 PDF，而不是复制模板原件。"""
    require_unlocked()
    job = args.get("job") if isinstance(args.get("job"), dict) else {}
    template_id = int(job.get("template_id") or 0)
    supplier_name = str(job.get("supplier_name") or "").strip()[:200]
    payee_name = str(job.get("payee_name") or "").strip()[:200]
    bank_name = str(job.get("bank_name") or "").strip()[:200]
    bank_account = re.sub(r"\s+", "", str(job.get("bank_account") or "").strip())[:100]
    line_items = job.get("line_items") if isinstance(job.get("line_items"), list) else []
    delivery_address = str(job.get("delivery_address") or "").strip()[:300]
    delivery_deadline = str(job.get("delivery_deadline") or "").strip()[:300]
    payment_terms = str(job.get("payment_terms") or "").strip()[:500]
    if template_id <= 0 or not supplier_name or not line_items:
        raise RuntimeError("合同生成指令不完整，请补充合同模板、供应商和商品明细。")
    if not bool(job.get("preserve_other_terms", True)):
        raise RuntimeError("当前仅支持保留模板其它条款的安全生成模式。")

    conn = connect_db(library_dir())
    template = conn.execute(
        "SELECT * FROM document_templates WHERE id = ? AND enabled = 1",
        (template_id,),
    ).fetchone()
    if not template or str(template["document_type"] or "") != "contract":
        conn.close()
        raise RuntimeError("指定的资源池模板不是合同模板。")
    template_text = str(template["template_text"] or template["content_text"] or "")
    if len(template_text) < 200:
        conn.close()
        raise RuntimeError("合同模板正文不完整，暂时无法保证原条款逐字保留。")

    support_texts: list[str] = []
    seen_ids: set[int] = set()
    for value in list(job.get("supporting_ids") or [])[:20]:
        try:
            support_id = int(value)
        except (TypeError, ValueError):
            continue
        if support_id <= 0 or support_id in seen_ids:
            continue
        seen_ids.add(support_id)
        row = conn.execute(
            "SELECT managed_path, original_path FROM documents WHERE id = ?", (support_id,)
        ).fetchone()
        if not row:
            continue
        path = Path(str(row["managed_path"] or row["original_path"] or ""))
        if path.exists():
            support_text = docintel.extract_text(path)
            if support_text:
                support_texts.append(support_text[:30000])
    now = datetime.now().isoformat(timespec="seconds")
    conn.execute(
        "UPDATE document_templates SET use_count = use_count + 1, last_used_at = ? WHERE id = ?",
        (now, template_id),
    )
    conn.commit()
    source_document_id = int(template["source_document_id"] or 0)
    template_name = str(template["name"] or "")
    conn.close()

    base_name = _safe_name(str(job.get("output_name") or "新合同"), 80) or "新合同"
    if not base_name.lower().endswith(".pdf"):
        base_name += ".pdf"
    out_dir = Path(tempfile.gettempdir()) / ("合同生成_" + time.strftime("%Y%m%d-%H%M%S"))
    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = out_dir / base_name
    info = contractgen.generate_contract_pdf(
        template_text=template_text,
        output_path=output_path,
        supplier_name=supplier_name,
        payee_name=payee_name,
        bank_name=bank_name,
        bank_account=bank_account,
        support_texts=support_texts,
        line_items=line_items,
        delivery_address=delivery_address,
        delivery_deadline=delivery_deadline,
        payment_terms=payment_terms,
    )
    if not output_path.exists() or output_path.stat().st_size < 1000:
        raise RuntimeError("合同文件生成失败。")
    return {
        "file_path": str(output_path),
        "dir": str(out_dir),
        "file_name": output_path.name,
        "template_id": template_id,
        "template_name": template_name,
        "source_document_id": source_document_id,
        "size_bytes": output_path.stat().st_size,
        "info": info,
    }


def command_generate_document(args: dict[str, Any]) -> dict[str, Any]:
    """生成 AI 综合起草文书；合同固定落地 PDF，其它文书生成 DOCX。"""
    require_unlocked()
    job = args.get("job") if isinstance(args.get("job"), dict) else {}
    title = str(job.get("title") or "新文书").strip()[:100]
    content = str(job.get("content") or "").strip()[:80000]
    document_type = str(job.get("document_type") or "other").strip().lower()
    if len(content) < 20:
        raise RuntimeError("AI 没有给出完整文书正文，暂不生成文件。")
    template_id = int(job.get("template_id") or 0)
    template_name = ""
    if template_id > 0:
        conn = connect_db(library_dir())
        row = conn.execute(
            "SELECT id, name FROM document_templates WHERE id = ? AND enabled = 1", (template_id,)
        ).fetchone()
        if not row:
            conn.close()
            raise RuntimeError("文书模板不存在或已停用。")
        template_name = str(row["name"] or "")
        conn.execute(
            "UPDATE document_templates SET use_count = use_count + 1, last_used_at = ? WHERE id = ?",
            (datetime.now().isoformat(timespec="seconds"), template_id),
        )
        conn.commit()
        conn.close()

    base_name = _safe_name(str(job.get("output_name") or title), 80) or "新文书"
    is_contract = document_type in {"contract", "agreement", "purchase_contract", "sales_contract"} or "合同" in title
    base_name = re.sub(r"\.(?:pdf|docx)$", "", base_name, flags=re.I)
    base_name += ".pdf" if is_contract else ".docx"
    out_dir = Path(tempfile.gettempdir()) / (
        ("合同生成_" if is_contract else "文书生成_") + time.strftime("%Y%m%d-%H%M%S")
    )
    output_path = out_dir / base_name
    info = None
    if is_contract:
        info = contractgen.generate_text_pdf(title, content, output_path)
    else:
        documentgen.generate_docx(title, content, output_path)
    return {
        "file_path": str(output_path),
        "dir": str(out_dir),
        "file_name": output_path.name,
        "template_id": template_id,
        "template_name": template_name,
        "size_bytes": output_path.stat().st_size,
        "format": "pdf" if is_contract else "docx",
        "info": info,
    }


def _training_note_payload(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": int(row["id"]),
        "title": str(row["title"] or ""),
        "trigger_keywords": str(row["trigger_keywords"] or ""),
        "instruction": str(row["instruction"] or ""),
        "enabled": bool(row["enabled"]),
        "use_count": int(row["use_count"] or 0),
        "last_used_at": str(row["last_used_at"] or ""),
        "created_at": str(row["created_at"] or ""),
        "updated_at": str(row["updated_at"] or ""),
    }


def _matching_training_notes(conn: sqlite3.Connection, message: str) -> list[dict[str, Any]]:
    rows = conn.execute(
        "SELECT * FROM assistant_training_notes WHERE enabled = 1 ORDER BY updated_at DESC LIMIT 300"
    ).fetchall()
    query = message.casefold()
    ranked: list[tuple[int, sqlite3.Row]] = []
    for row in rows:
        title = str(row["title"] or "").strip().casefold()
        keywords = [
            token.strip().casefold()
            for token in re.split(r"[,，、;；\n]+", str(row["trigger_keywords"] or ""))
            if token.strip()
        ]
        score = sum(5 for token in keywords if token in query)
        if title and title in query:
            score += 8
        # 两个字以上的标题片段也参与召回，避免用户必须逐字重复场景名。
        score += sum(1 for token in re.split(r"[\s/\-_]+", title) if len(token) >= 2 and token in query)
        if score > 0:
            ranked.append((score, row))
    ranked.sort(key=lambda item: (item[0], int(item[1]["id"])), reverse=True)
    matched = [row for _, row in ranked[:8]]
    if matched:
        now = datetime.now().isoformat(timespec="seconds")
        conn.executemany(
            "UPDATE assistant_training_notes SET use_count = use_count + 1, last_used_at = ? WHERE id = ?",
            [(now, int(row["id"])) for row in matched],
        )
        conn.commit()
    return [_training_note_payload(row) for row in matched]


def command_assistant_notes_list(_: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    conn = connect_db(library_dir())
    rows = conn.execute("SELECT * FROM assistant_training_notes ORDER BY updated_at DESC, id DESC").fetchall()
    conn.close()
    return {"notes": [_training_note_payload(row) for row in rows]}


def command_assistant_note_save(args: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    note_id = int(args.get("id") or 0)
    title = str(args.get("title") or "").strip()[:100]
    keywords = str(args.get("trigger_keywords") or "").strip()[:500]
    instruction = str(args.get("instruction") or "").strip()[:12000]
    enabled = 1 if bool(args.get("enabled", True)) else 0
    if not title:
        raise RuntimeError("请填写培训场景名称。")
    if not keywords:
        raise RuntimeError("请填写触发关键词，多个关键词用逗号分隔。")
    if not instruction:
        raise RuntimeError("请填写要教给档案秘书的处理方法。")

    conn = connect_db(library_dir())
    now = datetime.now().isoformat(timespec="seconds")
    if note_id > 0:
        exists = conn.execute("SELECT id FROM assistant_training_notes WHERE id = ?", (note_id,)).fetchone()
        if not exists:
            conn.close()
            raise RuntimeError("培训笔记不存在。")
        conn.execute(
            """UPDATE assistant_training_notes
               SET title = ?, trigger_keywords = ?, instruction = ?, enabled = ?, updated_at = ?
               WHERE id = ?""",
            (title, keywords, instruction, enabled, now, note_id),
        )
    else:
        cur = conn.execute(
            """INSERT INTO assistant_training_notes
               (title, trigger_keywords, instruction, enabled, use_count, created_at, updated_at)
               VALUES (?, ?, ?, ?, 0, ?, ?)""",
            (title, keywords, instruction, enabled, now, now),
        )
        note_id = int(cur.lastrowid)
    conn.commit()
    row = conn.execute("SELECT * FROM assistant_training_notes WHERE id = ?", (note_id,)).fetchone()
    conn.close()
    return {"note": _training_note_payload(row)}


def command_assistant_note_delete(args: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    note_id = int(args.get("id") or 0)
    if note_id <= 0:
        raise RuntimeError("缺少培训笔记 id。")
    conn = connect_db(library_dir())
    conn.execute("DELETE FROM assistant_training_notes WHERE id = ?", (note_id,))
    conn.commit()
    conn.close()
    return {"deleted": True, "id": note_id}


CHAT_FOLDER_NAME = "AI秘书聊天记录"


def _assistant_chat_dir() -> Path:
    target = library_dir() / CHAT_FOLDER_NAME
    target.mkdir(parents=True, exist_ok=True)
    return target


def _assistant_chat_path(conversation_id: str) -> Path:
    safe_id = str(conversation_id or "").strip().lower()
    if not re.fullmatch(r"[0-9a-f]{32}", safe_id):
        raise RuntimeError("聊天记录编号无效。")
    return _assistant_chat_dir() / f"{safe_id}.json"


def _read_assistant_conversation(path: Path) -> dict[str, Any] | None:
    try:
        data = json.loads(path.read_text("utf-8"))
        return data if isinstance(data, dict) else None
    except Exception:
        return None


def _clean_saved_message(item: Any) -> dict[str, Any] | None:
    if not isinstance(item, dict):
        return None
    role = "ai" if str(item.get("role") or "") in {"ai", "assistant"} else "user"
    text = str(item.get("text") or item.get("content") or "").strip()[:30000]
    if not text:
        return None
    message: dict[str, Any] = {
        "role": role,
        "text": text,
        "created_at": str(item.get("created_at") or datetime.now().isoformat(timespec="seconds"))[:40],
    }
    materials = item.get("materials") if isinstance(item.get("materials"), list) else []
    if materials:
        message["materials"] = [
            {
                "id": int(material.get("id") or 0),
                "file_name": str(material.get("file_name") or "")[:255],
                "type_label": str(material.get("type_label") or "")[:100],
                "company": str(material.get("company") or "")[:200],
                "reason": str(material.get("reason") or "")[:200],
                "picked": bool(material.get("picked", True)),
            }
            for material in materials[:100]
            if isinstance(material, dict) and int(material.get("id") or 0) > 0
        ]
    generated = item.get("generatedDocument") if isinstance(item.get("generatedDocument"), dict) else item.get("generated_document")
    if isinstance(generated, dict):
        message["generated_document"] = {
            "file_path": str(generated.get("file_path") or "")[:1000],
            "dir": str(generated.get("dir") or "")[:1000],
            "file_name": str(generated.get("file_name") or "")[:255],
            "template_name": str(generated.get("template_name") or "")[:100],
            "info": generated.get("info") if isinstance(generated.get("info"), dict) else {},
        }
    generated_sheet = item.get("generatedSpreadsheet") if isinstance(item.get("generatedSpreadsheet"), dict) else item.get("generated_spreadsheet")
    if isinstance(generated_sheet, dict) and generated_sheet.get("file_path"):
        message["generated_spreadsheet"] = {
            "file_path": str(generated_sheet.get("file_path") or "")[:1000],
            "dir": str(generated_sheet.get("dir") or "")[:1000],
            "file_name": str(generated_sheet.get("file_name") or "")[:255],
            "rows": max(0, int(generated_sheet.get("rows") or 0)),
            "missing": generated_sheet.get("missing") if isinstance(generated_sheet.get("missing"), dict) else {},
        }
    managed_task = item.get("managedTask") if isinstance(item.get("managedTask"), dict) else item.get("managed_task")
    if isinstance(managed_task, dict) and managed_task.get("id"):
        message["managed_task"] = {
            "id": str(managed_task.get("id") or "")[:40],
            "title": str(managed_task.get("title") or "表格托管事项")[:100],
            "status": str(managed_task.get("status") or "received")[:30],
            "files": [
                {"original_name": str(file.get("original_name") or "")[:255]}
                for file in list(managed_task.get("files") or [])[:20]
                if isinstance(file, dict)
            ],
        }
    contract_context = _contract_context_payload(item.get("contractContext") or item.get("contract_context"))
    if contract_context:
        message["contract_context"] = contract_context
    for key in ("organize", "useWatermark"):
        if key in item:
            message[key] = bool(item.get(key))
    if item.get("watermarkText"):
        message["watermarkText"] = str(item.get("watermarkText"))[:60]
    if role == "ai":
        message["rateable"] = bool(item.get("rateable", True))
        quick_options = []
        for option in list(item.get("quickOptions") or item.get("quick_options") or [])[:4]:
            if isinstance(option, dict):
                label = str(option.get("label") or option.get("message") or "").strip()[:80]
                option_message = str(option.get("message") or option.get("label") or "").strip()[:2000]
                if label and option_message:
                    quick_options.append({"label": label, "message": option_message})
            else:
                value = str(option or "").strip()[:200]
                if value:
                    quick_options.append(value)
        if quick_options:
            message["quick_options"] = quick_options
        try:
            rating = int(item.get("rating") or 0)
        except (TypeError, ValueError):
            rating = 0
        if 1 <= rating <= 5:
            message["rating"] = rating
            message["rating_feedback"] = str(
                item.get("ratingFeedback") or item.get("rating_feedback") or ""
            ).strip()[:2000]
            message["rated_at"] = str(
                item.get("ratedAt") or item.get("rated_at") or datetime.now().isoformat(timespec="seconds")
            )[:40]
    return message


def command_assistant_conversations_list(_: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    target = _assistant_chat_dir()
    conversations = []
    for path in target.glob("*.json"):
        data = _read_assistant_conversation(path)
        if not data:
            continue
        messages = data.get("messages") if isinstance(data.get("messages"), list) else []
        preview = ""
        for item in reversed(messages):
            if isinstance(item, dict) and str(item.get("text") or "").strip():
                preview = str(item.get("text") or "").strip().replace("\n", " ")[:120]
                break
        conversations.append({
            "id": str(data.get("id") or path.stem),
            "title": str(data.get("title") or "未命名对话")[:100],
            "created_at": str(data.get("created_at") or ""),
            "updated_at": str(data.get("updated_at") or ""),
            "message_count": len(messages),
            "preview": preview,
        })
    conversations.sort(key=lambda item: item["updated_at"], reverse=True)
    return {"conversations": conversations[:1000], "dir": str(target)}


def command_assistant_conversation_get(args: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    path = _assistant_chat_path(str(args.get("id") or ""))
    data = _read_assistant_conversation(path) if path.exists() else None
    if not data:
        raise RuntimeError("聊天记录不存在或已经删除。")
    return {"conversation": data, "dir": str(path.parent)}


def command_assistant_conversation_save(args: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    conversation_id = str(args.get("id") or "").strip().lower()
    if not conversation_id:
        conversation_id = uuid.uuid4().hex
    path = _assistant_chat_path(conversation_id)
    existing = _read_assistant_conversation(path) if path.exists() else None
    raw_messages = args.get("messages") if isinstance(args.get("messages"), list) else []
    messages = []
    for raw in raw_messages[-500:]:
        cleaned = _clean_saved_message(raw)
        if cleaned:
            messages.append(cleaned)
    if not messages:
        raise RuntimeError("没有可保存的聊天内容。")
    title = str(args.get("title") or (existing or {}).get("title") or "").strip()[:100]
    if not title or title == "新对话":
        first_user = next((m["text"] for m in messages if m["role"] == "user"), "新对话")
        title = re.sub(r"\s+", " ", first_user).strip()[:32] or "新对话"
    now = datetime.now().isoformat(timespec="seconds")
    data = {
        "version": 1,
        "id": conversation_id,
        "title": title,
        "created_at": str((existing or {}).get("created_at") or now),
        "updated_at": now,
        "messages": messages,
    }
    temp_path = path.with_suffix(".tmp")
    temp_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), "utf-8")
    os.replace(temp_path, path)
    return {"conversation": data, "file_path": str(path), "dir": str(path.parent)}


def command_assistant_conversation_rename(args: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    path = _assistant_chat_path(str(args.get("id") or ""))
    data = _read_assistant_conversation(path) if path.exists() else None
    if not data:
        raise RuntimeError("聊天记录不存在或已经删除。")
    title = str(args.get("title") or "").strip()[:100]
    if not title:
        raise RuntimeError("对话名称不能为空。")
    data["title"] = title
    data["updated_at"] = datetime.now().isoformat(timespec="seconds")
    temp_path = path.with_suffix(".tmp")
    temp_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), "utf-8")
    os.replace(temp_path, path)
    return {"conversation": data}


def command_assistant_conversation_delete(args: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    path = _assistant_chat_path(str(args.get("id") or ""))
    if path.exists():
        path.unlink()
    return {"deleted": True, "id": path.stem}


def command_assistant_training_export(_: dict[str, Any]) -> dict[str, Any]:
    """把已评分问答导出为本地 JSONL；资料只记录引用，不上传原文件。"""
    require_unlocked()
    records: list[dict[str, Any]] = []
    for path in sorted(_assistant_chat_dir().glob("*.json")):
        conversation = _read_assistant_conversation(path)
        if not conversation:
            continue
        messages = conversation.get("messages") if isinstance(conversation.get("messages"), list) else []
        for index, message in enumerate(messages):
            if not isinstance(message, dict) or message.get("role") != "ai":
                continue
            try:
                rating = int(message.get("rating") or 0)
            except (TypeError, ValueError):
                rating = 0
            if not 1 <= rating <= 5:
                continue
            previous_user = next(
                (
                    str(candidate.get("text") or "")
                    for candidate in reversed(messages[:index])
                    if isinstance(candidate, dict) and candidate.get("role") == "user"
                ),
                "",
            )
            if not previous_user:
                continue
            context = [
                {"role": str(candidate.get("role") or ""), "text": str(candidate.get("text") or "")}
                for candidate in messages[max(0, index - 6):index]
                if isinstance(candidate, dict) and str(candidate.get("text") or "").strip()
            ]
            materials = message.get("materials") if isinstance(message.get("materials"), list) else []
            contract_context = message.get("contract_context") if isinstance(message.get("contract_context"), dict) else {}
            document_refs = sorted({
                int(material.get("id") or 0)
                for material in materials
                if isinstance(material, dict) and int(material.get("id") or 0) > 0
            } | ({int(contract_context.get("source_document_id") or 0)} if int(contract_context.get("source_document_id") or 0) > 0 else set()))
            records.append({
                "schema_version": 1,
                "conversation_id": str(conversation.get("id") or path.stem),
                "conversation_title": str(conversation.get("title") or ""),
                "question": previous_user,
                "answer": str(message.get("text") or ""),
                "rating": rating,
                "feedback": str(message.get("rating_feedback") or ""),
                "rated_at": str(message.get("rated_at") or ""),
                "context": context,
                "document_refs": document_refs,
                "materials": materials,
                "generated_document": message.get("generated_document") if isinstance(message.get("generated_document"), dict) else None,
            })
    if not records:
        raise RuntimeError("还没有已评分的 AI 回复，请先给回答打分。")
    export_dir = library_dir() / "AI秘书训练数据"
    export_dir.mkdir(parents=True, exist_ok=True)
    output_path = export_dir / ("AI秘书训练数据_" + time.strftime("%Y%m%d-%H%M%S") + ".jsonl")
    payload = "\n".join(json.dumps(record, ensure_ascii=False) for record in records) + "\n"
    output_path.write_text(payload, "utf-8")
    return {
        "file_path": str(output_path),
        "dir": str(export_dir),
        "file_name": output_path.name,
        "count": len(records),
    }


def command_templates_list(_: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    conn = connect_db(library_dir())
    rows = conn.execute(
        "SELECT * FROM document_templates ORDER BY enabled DESC, updated_at DESC, id DESC"
    ).fetchall()
    conn.close()
    return {
        "templates": [templatepool.payload(row) for row in rows],
        "active": sum(1 for row in rows if bool(row["enabled"])),
    }


def command_templates_rebuild(_: dict[str, Any]) -> dict[str, Any]:
    """从现有已识别资料补建资源池；适合升级后的第一次使用。"""
    require_unlocked()
    conn = connect_db(library_dir())
    rows = conn.execute(
        "SELECT id, review_status FROM documents WHERE recognized_at IS NOT NULL ORDER BY id ASC"
    ).fetchall()
    created = 0
    skipped = 0
    for row in rows:
        try:
            result = templatepool.upsert_from_document(
                conn,
                int(row["id"]),
                active=str(row["review_status"] or "") == "confirmed",
            )
            if result:
                created += 1
            else:
                skipped += 1
        except Exception:
            skipped += 1
    conn.close()
    return {"processed": len(rows), "templates": created, "skipped": skipped}


def command_template_save(args: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    template_id = int(args.get("id") or 0)
    if template_id <= 0:
        raise RuntimeError("缺少模板 id。")
    name = str(args.get("name") or "").strip()[:100]
    content = str(args.get("template_text") or "").strip()[:120000]
    if not name:
        raise RuntimeError("请填写模板名称。")
    if len(content) < 20:
        raise RuntimeError("模板正文过短。")
    enabled = 1 if bool(args.get("enabled", True)) else 0
    conn = connect_db(library_dir())
    exists = conn.execute("SELECT id FROM document_templates WHERE id = ?", (template_id,)).fetchone()
    if not exists:
        conn.close()
        raise RuntimeError("模板不存在。")
    conn.execute(
        """UPDATE document_templates SET name = ?, template_text = ?, enabled = ?, status = 'active',
           updated_at = ? WHERE id = ?""",
        (name, content, enabled, datetime.now().isoformat(timespec="seconds"), template_id),
    )
    conn.commit()
    result = templatepool.get_template(conn, template_id)
    conn.close()
    return {"template": result}


def command_template_delete(args: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    template_id = int(args.get("id") or 0)
    if template_id <= 0:
        raise RuntimeError("缺少模板 id。")
    conn = connect_db(library_dir())
    conn.execute("DELETE FROM document_templates WHERE id = ?", (template_id,))
    conn.commit()
    conn.close()
    return {"deleted": True, "id": template_id}


def command_gather_to_folder(args: dict[str, Any]) -> dict[str, Any]:
    """把选中的资料复制到一个临时文件夹（用规范文件名），返回文件夹路径，前端打开它。
    args: {ids:[...]}"""
    require_unlocked()
    ids = args.get("ids") or []
    use_watermark = bool(args.get("use_watermark"))
    watermark_text = str(args.get("watermark_text") or "").strip()[:60]
    if use_watermark and not watermark_text:
        watermark_text = "仅供本次业务办理使用，再次复印无效"
    conn = connect_db(library_dir())
    picks = []
    for did in ids:
        try:
            r = conn.execute("SELECT managed_path, original_path, file_name FROM documents WHERE id = ?", (int(did),)).fetchone()
        except Exception:
            continue
        if not r:
            continue
        p = str(r["managed_path"] or "")
        if not p or not Path(p).exists():
            p = str(r["original_path"] or "")
        if p and Path(p).exists():
            picks.append((Path(p), str(r["file_name"] or "")))
    conn.close()
    if not picks:
        raise RuntimeError("没有可整理的文件（可能原文件已移动或删除）。")
    out_dir = Path(tempfile.gettempdir()) / ("资料整理_" + time.strftime("%Y%m%d-%H%M%S"))
    out_dir.mkdir(parents=True, exist_ok=True)
    count = 0
    watermarked = 0
    watermark_skipped: list[str] = []
    for src, name in picks:
        target_name = name or src.name
        dest = out_dir / target_name
        i = 2
        while dest.exists():
            dest = out_dir / f"{Path(target_name).stem}({i}){Path(target_name).suffix}"
            i += 1
        try:
            if use_watermark:
                applied, reason = _export_with_watermark(src, dest, watermark_text)
                if applied:
                    watermarked += 1
                elif reason:
                    watermark_skipped.append(f"{src.name}：{reason}")
            else:
                shutil.copy2(str(src), str(dest))
            count += 1
        except Exception as exc:
            watermark_skipped.append(f"{src.name}：{str(exc)[:100]}")
    return {
        "dir": str(out_dir),
        "count": count,
        "watermarked": watermarked,
        "watermark_text": watermark_text if use_watermark else "",
        "watermark_skipped": watermark_skipped[:20],
    }


def _watermark_font(size: int):
    from PIL import ImageFont

    candidates = [
        Path(os.environ.get("WINDIR", "C:/Windows")) / "Fonts" / "msyh.ttc",
        Path(os.environ.get("WINDIR", "C:/Windows")) / "Fonts" / "simhei.ttf",
        Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    for path in candidates:
        if path.exists():
            try:
                return ImageFont.truetype(str(path), size)
            except Exception:
                continue
    return ImageFont.load_default()


def _watermark_overlay(width: int, height: int, text: str):
    """生成透明的斜向平铺水印层，供 PDF 和图片副本复用。"""
    from PIL import Image, ImageDraw

    width, height = max(320, width), max(320, height)
    font_size = max(22, min(52, round(min(width, height) * 0.055)))
    font = _watermark_font(font_size)
    probe = ImageDraw.Draw(Image.new("RGBA", (4, 4), (0, 0, 0, 0)))
    box = probe.textbbox((0, 0), text, font=font)
    tw, th = max(1, box[2] - box[0]), max(1, box[3] - box[1])
    tile = Image.new("RGBA", (tw + 80, th + 54), (0, 0, 0, 0))
    ImageDraw.Draw(tile).text((40, 20), text, font=font, fill=(92, 112, 103, 52))
    tile = tile.rotate(28, expand=True, resample=Image.Resampling.BICUBIC)

    overlay = Image.new("RGBA", (width, height), (0, 0, 0, 0))
    step_x = max(tile.width + 70, round(width * 0.42))
    step_y = max(tile.height + 45, round(height * 0.22))
    row = 0
    for y in range(-tile.height, height + tile.height, step_y):
        offset = -(step_x // 2) if row % 2 else 0
        for x in range(-tile.width + offset, width + tile.width, step_x):
            overlay.alpha_composite(tile, (x, y))
        row += 1
    return overlay


def _watermark_pdf(source: Path, dest: Path, text: str) -> None:
    import fitz

    doc = fitz.open(str(source))
    try:
        for page in doc:
            rect = page.rect
            scale = 1.15
            width, height = max(320, int(rect.width * scale)), max(320, int(rect.height * scale))
            overlay = _watermark_overlay(width, height, text)
            stream = io.BytesIO()
            overlay.save(stream, format="PNG", optimize=True)
            page.insert_image(rect, stream=stream.getvalue(), overlay=True, keep_proportion=False)
        doc.save(str(dest), garbage=4, deflate=True)
    finally:
        doc.close()


def _watermark_image(source: Path, dest: Path, text: str) -> None:
    from PIL import Image, ImageOps

    with Image.open(source) as image:
        image = ImageOps.exif_transpose(image).convert("RGBA")
        output = Image.alpha_composite(image, _watermark_overlay(image.width, image.height, text))
        suffix = dest.suffix.lower()
        if suffix in {".jpg", ".jpeg"}:
            output.convert("RGB").save(dest, quality=95, optimize=True)
        else:
            output.save(dest)


def _watermark_excel_header(source: Path, dest: Path, text: str) -> None:
    import openpyxl

    keep_vba = source.suffix.lower() == ".xlsm"
    workbook = openpyxl.load_workbook(str(source), keep_vba=keep_vba)
    try:
        for sheet in workbook.worksheets:
            sheet.oddHeader.center.text = "&KAAAAAA&18" + text
            sheet.oddHeader.center.size = 18
        workbook.save(str(dest))
    finally:
        workbook.close()


def _export_with_watermark(source: Path, dest: Path, text: str) -> tuple[bool, str]:
    """只处理导出副本。失败或暂不支持时复制原文件并返回原因。"""
    suffix = source.suffix.lower()
    try:
        if suffix == ".pdf":
            _watermark_pdf(source, dest, text)
            return True, ""
        if suffix in IMAGE_EXTENSIONS:
            _watermark_image(source, dest, text)
            return True, ""
        if suffix in {".xlsx", ".xlsm"}:
            _watermark_excel_header(source, dest, text)
            return True, ""
        shutil.copy2(str(source), str(dest))
        return False, "当前格式暂不支持写入水印，已按原文件导出"
    except Exception as exc:
        if not dest.exists():
            shutil.copy2(str(source), str(dest))
        return False, "水印写入失败，已按原文件导出（" + str(exc)[:80] + "）"


def _gather_cache_dirs() -> list[Path]:
    """临时目录里的资料整理及合同生成缓存文件夹。"""
    base = Path(tempfile.gettempdir())
    try:
        result = [d for d in base.glob("资料整理_*") if d.is_dir()]
        result.extend(d for d in base.glob("合同生成_*") if d.is_dir())
        result.extend(d for d in base.glob("文书生成_*") if d.is_dir())
        return result
    except Exception:
        return []


def command_cache_info(_: dict[str, Any]) -> dict[str, Any]:
    """整理缓存占用：文件夹个数 + 总字节。超 1G 提示用户清理。这些只是合同副本，删了不影响原合同。"""
    folders = _gather_cache_dirs()
    total = 0
    for d in folders:
        for f in d.rglob("*"):
            if f.is_file():
                try:
                    total += f.stat().st_size
                except Exception:
                    pass
    return {"count": len(folders), "bytes": total, "over_limit": total > 1024 ** 3}


def command_clear_cache(_: dict[str, Any]) -> dict[str, Any]:
    """一键清空所有「资料整理_*」缓存文件夹（只删整理出来的副本，不动资料库和原合同）。"""
    folders = _gather_cache_dirs()
    removed = 0
    freed = 0
    locked = 0
    for d in folders:
        try:
            size = sum(f.stat().st_size for f in d.rglob("*") if f.is_file())
        except Exception:
            size = 0
        shutil.rmtree(d, ignore_errors=True)
        if not d.exists():            # 真正删掉才算（被占用/打开的删不掉，不误报）
            removed += 1
            freed += size
        else:
            locked += 1
    return {"removed": removed, "freed_bytes": freed, "locked": locked}


def command_library_summary(_: dict[str, Any]) -> dict[str, Any]:
    require_unlocked()
    try:
        root = library_dir()
    except RuntimeError:
        return {"configured": False, "documents": 0, "pages": 0, "library_dir": ""}
    conn = connect_db(root)
    row = conn.execute("SELECT COUNT(*), COALESCE(SUM(page_count), 0) FROM documents").fetchone()
    recent = conn.execute(
        "SELECT file_name, extension, page_count, status, created_at FROM documents ORDER BY id DESC LIMIT 12"
    ).fetchall()
    conn.close()
    return {
        "configured": True,
        "library_dir": str(root),
        "documents": int(row[0] or 0),
        "pages": int(row[1] or 0),
        "recent": [
            {
                "name": item[0],
                "extension": item[1],
                "pages": item[2],
                "status": item[3],
                "created_at": item[4],
            }
            for item in recent
        ],
    }


COMMANDS = {
    "get_serial": command_get_serial,
    "register": command_register,
    "sync_quota": command_sync_quota,
    "get_state": command_get_state,
    "set_password": command_set_password,
    "verify_password": command_verify_password,
    "set_library_dir": command_set_library_dir,
    "start_knowledge_index": command_start_knowledge_index,
    "scan_roots": command_scan_roots,
    "scan_folder": command_scan_folder,
    "quick_scan": command_quick_scan,
    "quick_scan_list": command_quick_scan_list,
    "quick_scan_mark": command_quick_scan_mark,
    "quick_scan_delete": command_quick_scan_delete,
    "quick_scan_import": command_quick_scan_import,
    "import_files": command_import_files,
    "find_materials": command_find_materials,
    "assistant_chat": command_assistant_chat,
    "assistant_task_create": command_assistant_task_create,
    "assistant_tasks_list": command_assistant_tasks_list,
    "assistant_task_get": command_assistant_task_get,
    "assistant_task_archive": command_assistant_task_archive,
    "assistant_task_organize": command_assistant_task_organize,
    "generate_contract": command_generate_contract,
    "generate_document": command_generate_document,
    "assistant_notes_list": command_assistant_notes_list,
    "assistant_note_save": command_assistant_note_save,
    "assistant_note_delete": command_assistant_note_delete,
    "assistant_conversations_list": command_assistant_conversations_list,
    "assistant_conversation_get": command_assistant_conversation_get,
    "assistant_conversation_save": command_assistant_conversation_save,
    "assistant_conversation_rename": command_assistant_conversation_rename,
    "assistant_conversation_delete": command_assistant_conversation_delete,
    "assistant_training_export": command_assistant_training_export,
    "templates_list": command_templates_list,
    "templates_rebuild": command_templates_rebuild,
    "template_save": command_template_save,
    "template_delete": command_template_delete,
    "gather_to_folder": command_gather_to_folder,
    "cache_info": command_cache_info,
    "clear_cache": command_clear_cache,
    "analyze_document": command_analyze_document,
    "analyze_documents": command_analyze_documents,
    "cancel_recognition": command_cancel_recognition,
    "document_meta": command_document_meta,
    "unrecognized_documents": command_unrecognized_documents,
    "list_documents": command_list_documents,
    "get_document": command_get_document,
    "confirm_document": command_confirm_document,
    "delete_document": command_delete_document,
    "library_summary": command_library_summary,
}


def _execute_request(request_id: Any, cmd: str, args: dict[str, Any]) -> None:
    try:
        data = COMMANDS[cmd](args)
        emit({"id": request_id, "ok": True, "data": data})
    except Exception as exc:
        emit({"id": request_id, "ok": False, "error": str(exc)})


def main() -> None:
    # Electron 的 child.stdin 固定写 UTF-8；从二进制管道自行解码，不能依赖
    # Windows/PyInstaller 为 sys.stdin 选择的区域编码。
    input_stream = getattr(sys.stdin, "buffer", sys.stdin)
    for raw_line in input_stream:
        line = raw_line.decode("utf-8-sig") if isinstance(raw_line, bytes) else raw_line
        line = line.strip()
        if not line:
            continue
        try:
            message = json.loads(line)
            request_id = message.get("id")
            cmd = str(message.get("cmd") or "")
            args = message.get("args") or {}
            if cmd not in COMMANDS:
                raise RuntimeError(f"未知命令：{cmd}")
            # 批量识别放到后台线程，主循环才能继续接收“取消识别”命令。
            if cmd == "analyze_documents":
                threading.Thread(
                    target=_execute_request,
                    args=(request_id, cmd, args),
                    name=f"aidoc-request-{request_id}",
                    daemon=False,
                ).start()
            else:
                _execute_request(request_id, cmd, args)
        except Exception as exc:
            emit({"id": locals().get("request_id"), "ok": False, "error": str(exc)})


if __name__ == "__main__":
    main()
