"""Excel 模板生成 + 数据读取"""

from pathlib import Path
from typing import Iterator

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


def collect_columns(steps: list) -> list[str]:
    """从整理后的步骤里提取所有 Excel 列名（有序、去重）。"""
    cols = []
    seen = set()
    for s in steps:
        if not s.get("selected", True):
            continue
        col = (s.get("excel_column") or "").strip()
        if col and col not in seen:
            cols.append(col)
            seen.add(col)
    return cols


def collect_columns_from_dsl(dsl_obj: dict) -> list[str]:
    """
    从 AI 生成的 DSL 反向扫所有 from_excel 字段。
    用途：AI 可能"发明"新列名（如把 .el-upload 自动转成 upload_folder_to_library
          并指定 from_excel="商品图片目录"），录制步骤里没绑定这个列，必须从 DSL 补回。
    """
    cols = []
    seen = set()
    if not isinstance(dsl_obj, dict):
        return cols
    for a in dsl_obj.get("actions", []):
        if not isinstance(a, dict):
            continue
        col = (a.get("from_excel") or "").strip()
        if col and col not in seen:
            cols.append(col)
            seen.add(col)
    return cols


def merge_columns(steps_cols: list[str], dsl_cols: list[str]) -> list[str]:
    """合并两个来源的列名：录制时绑定的优先，AI 新加的追加在后。"""
    merged = list(steps_cols)
    seen = set(steps_cols)
    for c in dsl_cols:
        if c not in seen:
            merged.append(c)
            seen.add(c)
    return merged


def collect_sample_row(steps: list, cols: list[str]) -> dict:
    """根据步骤里的录制原值，组合出一行示例数据。

    优先用 value（input/fill 类的录制值），
    没有则用 text（select_option/click 类的可见文本），
    upload 类型给一个友好的文件夹路径占位。
    """
    row = {}
    upload_cols = set()  # 哪些列是上传步骤的
    for s in steps:
        if not s.get("selected", True):
            continue
        col = (s.get("excel_column") or "").strip()
        if not col or col not in cols or col in row:
            continue
        if s.get("action_type") == "upload":
            upload_cols.add(col)
            # 上传列：录制时没值，给一个占位提示用户填什么
            row[col] = r"D:\示例\填文件或文件夹路径"
            continue
        val = s.get("value")
        if val is None or str(val).strip() == "":
            val = s.get("text", "")
        row[col] = "" if val is None else str(val)
    return row


def is_upload_col(col_name: str) -> bool:
    """从列名识别是否为上传列（用于 DSL 反推时的兜底）"""
    if not col_name:
        return False
    keywords = ("文件夹", "目录", "图片", "附件", "路径", "文件", "_folder", "_dir")
    return any(k in col_name for k in keywords)


def generate_template(path: Path, cols: list[str], sample_row: dict | None = None):
    """生成 Excel 模板。

    布局：
      第 1 行：表头（绿色）
      第 2 行：录制时的原始示例数据（黄色背景，提示"录制时填的"）
      第 3 行起：用户填入批量数据
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"

    head_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="16A34A")
    sample_font = Font(name="Microsoft YaHei", size=10, italic=True, color="78350F")
    sample_fill = PatternFill("solid", fgColor="FEF3C7")  # 浅黄
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 第 1 行：表头
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = center
        # 列宽根据列名长度 + 示例值长度自适应
        sample_text = (sample_row or {}).get(c, "")
        width = max(14, max(len(str(c)) * 2 + 2, len(str(sample_text)) + 4))
        ws.column_dimensions[cell.column_letter].width = min(width, 32)

    # 第 2 行：示例数据（黄色高亮 + 提示）
    if sample_row:
        for i, c in enumerate(cols, 1):
            val = sample_row.get(c, "")
            cell = ws.cell(row=2, column=i, value=val)
            cell.font = sample_font
            cell.fill = sample_fill
            cell.alignment = left

    # 冻结首行（滚动时表头不动）
    ws.freeze_panes = "A2"

    # 说明 sheet
    ws2 = wb.create_sheet("使用说明")
    ws2["A1"] = "📋 使用说明"
    ws2["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True, color="16A34A")
    notes = [
        "",
        "📊 「数据」工作表说明",
        "",
        "  • 第 1 行（绿色表头）：列名，严格匹配不要修改",
        "  • 第 2 行（黄色示例）：录制时你实际填/选的值，**会作为第 1 次循环的数据被执行**",
        "  • 第 3 行起：你要批量执行的新数据，每行 = 1 次循环",
        "  • 空行 = 结束循环",
        "",
        "💡 对于「下拉菜单」列：",
        "  示例值就是录制时选的选项文字，照样填写即可",
        "  例如示例显示「集采不含运」，你后续每行填「集采不含运」或「集采含运」",
        "",
        "📁 对于「上传文件」列（列名含「文件夹」/「目录」/「图片」）：",
        "  支持两种填法：",
        "    • 填单个文件路径：D:\\图片\\商品1.jpg",
        "    • 填整个文件夹路径：D:\\商品图片\\sku001（自动上传里面所有文件）",
        "  支持的扩展名：图片(jpg/png/webp...)、文档(pdf/doc/docx)、表格(xls/xlsx)、视频(mp4)",
        "",
        "⚠️ 注意事项",
        "  • 不要修改第 1 行表头的列名",
        "  • 留空的单元格 = 该步骤跳过/用默认值",
        "  • 软件会从第 2 行（含示例）开始循环",
        "  • 不想保留示例行的话，删除第 2 行即可",
    ]
    for i, n in enumerate(notes, 2):
        cell = ws2.cell(row=i, column=1, value=n)
        cell.font = Font(name="Microsoft YaHei", size=11)
    ws2.column_dimensions["A"].width = 70

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def read_rows(path: Path) -> Iterator[dict]:
    """逐行读 Excel 数据。

    连续遇到多行空行才判定数据结束；中间零星的一两行空行会**跳过而非截断**，
    避免用户表格中间不小心空一行就把后面几十行数据静默丢掉。
    """
    # read_only 模式必须显式 close，否则 Windows 上文件句柄不释放，
    # 跑完流程后用户删不掉/改不了这个 Excel。用 try/finally 兜住。
    wb = load_workbook(str(path), data_only=True, read_only=True)
    try:
        ws = wb["数据"] if "数据" in wb.sheetnames else wb.active

        rows = ws.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            return
        headers = [str(h).strip() if h is not None else "" for h in headers]

        MAX_BLANK_STREAK = 5  # 连续 5 行全空才停（容忍中间偶尔空一行）
        blank_streak = 0
        for r in rows:
            if not r or all(v is None or str(v).strip() == "" for v in r):
                blank_streak += 1
                if blank_streak >= MAX_BLANK_STREAK:
                    break
                continue
            blank_streak = 0
            d = {}
            for h, v in zip(headers, r):
                if not h:
                    continue
                d[h] = "" if v is None else (str(v) if not isinstance(v, str) else v)
            if any(d.values()):
                yield d
    finally:
        wb.close()


def count_rows(path: Path) -> int:
    """统计 Excel 有效数据行数。"""
    return sum(1 for _ in read_rows(path))
